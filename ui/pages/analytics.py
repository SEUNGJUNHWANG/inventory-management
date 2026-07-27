"""
재고관리 시스템 - 분석 대시보드
탭 1: 월별 매출·원가·마진 분석
탭 2: BOM 변경이력
탭 3: 단가 변경이력
탭 4: 부품 입출고 요약
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
from datetime import datetime, timedelta
from collections import defaultdict
from core.constants import COLORS, FONT_FAMILY, FONT_SIZES


# ──────────────────────────────────────────────────────────────────────────────
# 공통 헬퍼
# ──────────────────────────────────────────────────────────────────────────────

def _만원(n):
    """숫자를 '1,234만원' 문자열로 변환"""
    if n == 0:
        return "0원"
    if abs(n) >= 10000:
        return f"{n / 10000:,.1f}만원"
    return f"{int(n):,}원"


def _pct(n):
    return f"{n:.1f}%"


class AnalyticsDashboard:
    """분석 대시보드 페이지 (탭 4개)"""

    def __init__(self, app):
        self.app = app
        # 탭별 로드 여부 (lazy load)
        self._tab_loaded = {0: False, 1: False, 2: False, 3: False}
        # 탭 1 상태
        self._sales_year = datetime.now().year
        self._sales_data = []
        # 탭 2 상태
        self._bom_log_data = []
        # 탭 3 상태
        self._price_log_data = []
        # 탭 4 상태
        self._io_year = datetime.now().year
        self._io_data = []

    # ══════════════════════════════════════════════════════════════════════════
    # render
    # ══════════════════════════════════════════════════════════════════════════
    def render(self):
        # 페이지 재진입 시 lazy-load 상태 초기화 (이전 render의 잔류 상태 제거)
        self._tab_loaded = {0: False, 1: False, 2: False, 3: False}
        scroll_frame = self.app._create_scrollable_frame()

        # 헤더
        hdr = tk.Frame(scroll_frame, bg=COLORS["bg"])
        hdr.pack(fill=tk.X, padx=5, pady=(0, 8))
        tk.Label(hdr, text="📈 분석 대시보드",
                 bg=COLORS["bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["title"], "bold")).pack(side=tk.LEFT)

        # Notebook
        style = ttk.Style()
        style.configure("Analytics.TNotebook.Tab",
                        font=(FONT_FAMILY, FONT_SIZES["small"], "bold"),
                        padding=[14, 6])
        nb = ttk.Notebook(scroll_frame, style="Analytics.TNotebook")
        nb.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 10))

        # 탭 프레임 생성
        self._tab_frames = []
        tab_names = [
            "📊 월별 매출·마진",
            "📋 BOM 변경이력",
            "💰 단가 변경이력",
            "🔄 부품 입출고 요약",
        ]
        for name in tab_names:
            f = tk.Frame(nb, bg=COLORS["bg"])
            nb.add(f, text=name)
            self._tab_frames.append(f)

        # 탭 전환 이벤트 → lazy load
        nb.bind("<<NotebookTabChanged>>",
                lambda e: self._on_tab_change(nb.index(nb.select())))

        # 첫 탭 바로 로드
        self._load_tab(0)

    # ══════════════════════════════════════════════════════════════════════════
    # 탭 전환 / lazy load
    # ══════════════════════════════════════════════════════════════════════════
    def _on_tab_change(self, idx):
        if not self._tab_loaded.get(idx):
            self._load_tab(idx)

    def _load_tab(self, idx):
        self._tab_loaded[idx] = True
        builders = [
            self._build_sales_tab,
            self._build_bom_log_tab,
            self._build_price_log_tab,
            self._build_io_tab,
        ]
        builders[idx](self._tab_frames[idx])

    # ══════════════════════════════════════════════════════════════════════════
    # 탭 1 — 월별 매출·마진 분석
    # ══════════════════════════════════════════════════════════════════════════
    def _build_sales_tab(self, parent):
        # ── 필터 바 ──────────────────────────────────────────────────────────
        fbar = tk.Frame(parent, bg=COLORS["bg"])
        fbar.pack(fill=tk.X, padx=10, pady=8)

        tk.Label(fbar, text="조회 연도:", bg=COLORS["bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)

        years = [str(y) for y in range(datetime.now().year, datetime.now().year - 5, -1)]
        self._sales_year_var = tk.StringVar(value=str(self._sales_year))
        yr_cb = ttk.Combobox(fbar, textvariable=self._sales_year_var,
                             values=years, width=8, state="readonly")
        yr_cb.pack(side=tk.LEFT, padx=(4, 10))

        tk.Button(fbar, text="🔄 조회", font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg=COLORS["primary"], fg="white", padx=10, pady=3,
                  cursor="hand2", relief="flat",
                  command=self._reload_sales).pack(side=tk.LEFT)

        self._sales_note = tk.Label(
            fbar,
            text="※ 원가는 현재 BOM 기준 추산값입니다.",
            bg=COLORS["bg"], fg=COLORS["text_secondary"],
            font=(FONT_FAMILY, FONT_SIZES["tiny"]))
        self._sales_note.pack(side=tk.LEFT, padx=12)

        # ── 요약 카드 4개 ────────────────────────────────────────────────────
        self._sales_card_frame = tk.Frame(parent, bg=COLORS["bg"])
        self._sales_card_frame.pack(fill=tk.X, padx=10, pady=(0, 8))

        # ── 차트 영역 ────────────────────────────────────────────────────────
        chart_outer = tk.Frame(parent, bg=COLORS["card_bg"],
                               highlightbackground=COLORS["border"],
                               highlightthickness=1)
        chart_outer.pack(fill=tk.X, padx=10, pady=(0, 8))

        tk.Label(chart_outer, text="월별 매출 현황",
                 bg=COLORS["card_bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")
                 ).pack(anchor="w", padx=15, pady=(10, 4))

        self._sales_chart_frame = tk.Frame(chart_outer, bg=COLORS["card_bg"])
        self._sales_chart_frame.pack(fill=tk.X, padx=10, pady=(0, 10))

        # ── 상세 테이블 ──────────────────────────────────────────────────────
        tbl_outer = tk.Frame(parent, bg=COLORS["card_bg"],
                             highlightbackground=COLORS["border"],
                             highlightthickness=1)
        tbl_outer.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        tk.Label(tbl_outer, text="월별 상세",
                 bg=COLORS["card_bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")
                 ).pack(anchor="w", padx=15, pady=(10, 4))

        cols = ("월", "건수", "수량", "판매대금", "원가(추산)", "마진금액", "마진율")
        self._sales_tree = ttk.Treeview(tbl_outer, columns=cols,
                                        show="headings", height=14)
        widths = {"월": 85, "건수": 60, "수량": 70,
                  "판매대금": 115, "원가(추산)": 115,
                  "마진금액": 115, "마진율": 80}
        anchors = {"월": "center", "건수": "center", "수량": "center",
                   "판매대금": "e", "원가(추산)": "e", "마진금액": "e", "마진율": "center"}
        for c in cols:
            self._sales_tree.heading(c, text=c)
            self._sales_tree.column(c, width=widths[c], anchor=anchors[c])

        self._sales_tree.tag_configure("pos", foreground="#16a34a")
        self._sales_tree.tag_configure("neg", foreground="#dc2626")

        sb = ttk.Scrollbar(tbl_outer, orient="vertical",
                           command=self._sales_tree.yview)
        self._sales_tree.configure(yscrollcommand=sb.set)
        self._sales_tree.pack(side=tk.LEFT, fill=tk.BOTH,
                              expand=True, padx=(10, 0), pady=(0, 10))
        sb.pack(side=tk.RIGHT, fill=tk.Y, pady=(0, 10), padx=(0, 5))

        # 첫 데이터 로드
        self._reload_sales()

    def _reload_sales(self):
        try:
            self._sales_year = int(self._sales_year_var.get())
        except Exception:
            pass

        def load():
            try:
                data = self.app.db.get_monthly_sales_with_margin(self._sales_year)
                self.app.root.after(0, lambda: self._render_sales(data))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err))

        threading.Thread(target=load, daemon=True).start()

    def _render_sales(self, data):
        self._sales_data = data

        # ── 요약 카드 ────────────────────────────────────────────────────────
        for w in self._sales_card_frame.winfo_children():
            w.destroy()

        # 이번달 또는 가장 최근 데이터
        now_ym = datetime.now().strftime("%Y-%m")
        recent = next((d for d in reversed(data) if d["월"] <= now_ym), None) or {}
        total_rev  = sum(d["판매대금"] for d in data)
        total_cost = sum(d["원가"]     for d in data)
        total_mgn  = sum(d["마진"]     for d in data)
        avg_rate   = (total_mgn / total_rev * 100) if total_rev > 0 else 0

        cards = [
            ("연간 판매대금",  _만원(total_rev),  COLORS["primary"]),
            ("연간 원가(추산)", _만원(total_cost), COLORS["warning"]),
            ("연간 마진금액",  _만원(total_mgn),  COLORS["success"] if total_mgn >= 0 else COLORS["danger"]),
            ("평균 마진율",    _pct(avg_rate),    COLORS["info"]),
        ]
        for label, val, color in cards:
            card = tk.Frame(self._sales_card_frame, bg="white",
                            highlightbackground=color, highlightthickness=2)
            card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=3)
            tk.Label(card, text=label, bg="white",
                     fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["small"])).pack(pady=(8, 2))
            tk.Label(card, text=val, bg="white", fg=color,
                     font=(FONT_FAMILY, FONT_SIZES["stat"], "bold")).pack(pady=(0, 8))

        # ── 차트 ─────────────────────────────────────────────────────────────
        for w in self._sales_chart_frame.winfo_children():
            w.destroy()
        self._draw_sales_chart(data)

        # ── 테이블 ───────────────────────────────────────────────────────────
        self._sales_tree.delete(*self._sales_tree.get_children())
        for d in reversed(data):
            tag = "pos" if d["마진"] >= 0 else "neg"
            self._sales_tree.insert("", "end", values=(
                d["월"],
                f"{d['건수']:,}",
                f"{d['수량']:,}",
                f"{d['판매대금']:,}원",
                f"{d['원가']:,}원",
                f"{d['마진']:,}원",
                _pct(d["마진율"]),
            ), tags=(tag,))

        # 합계 행
        if data:
            self._sales_tree.insert("", "end", values=(
                "합계",
                f"{sum(d['건수'] for d in data):,}",
                f"{sum(d['수량'] for d in data):,}",
                f"{total_rev:,}원",
                f"{total_cost:,}원",
                f"{total_mgn:,}원",
                _pct(avg_rate),
            ), tags=("pos" if total_mgn >= 0 else "neg",))

    def _draw_sales_chart(self, data):
        """Canvas 기반 그룹형 바 차트 (판매대금·원가·마진)"""
        if not data:
            tk.Label(self._sales_chart_frame, text="데이터 없음",
                     bg=COLORS["card_bg"], fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["small"])).pack(pady=30)
            return

        COL_W    = 90
        BAR_MAX  = 120
        CANVAS_H = BAR_MAX + 70
        COLORS_3 = ["#3b82f6", "#f59e0b", "#22c55e"]  # 판매대금, 원가, 마진
        LABELS_3 = ["판매대금", "원가", "마진"]

        # 최대값 기준 (판매대금 기준)
        max_val = max((d["판매대금"] for d in data), default=1) or 1

        for d in data:
            col_f = tk.Frame(self._sales_chart_frame, bg=COLORS["card_bg"])
            col_f.pack(side=tk.LEFT, expand=True)

            canvas = tk.Canvas(col_f, width=COL_W, height=CANVAS_H,
                               bg=COLORS["card_bg"], highlightthickness=0)
            canvas.pack()

            bar_vals = [d["판매대금"], d["원가"], d["마진"]]
            bar_w    = 16
            gap      = 2
            total_bw = len(bar_vals) * bar_w + (len(bar_vals) - 1) * gap
            x_start  = (COL_W - total_bw) // 2
            bot_y    = CANVAS_H - 30

            for bi, (val, color) in enumerate(zip(bar_vals, COLORS_3)):
                x0  = x_start + bi * (bar_w + gap)
                x1  = x0 + bar_w
                h   = max(2, int(BAR_MAX * abs(val) / max_val)) if val != 0 else 2
                top = bot_y - h
                fill = color if val >= 0 else "#fca5a5"
                canvas.create_rectangle(x0, top, x1, bot_y,
                                        fill=fill, outline="")

            # 월 라벨
            canvas.create_text(COL_W // 2, CANVAS_H - 14,
                                text=d["월"][5:] + "월",
                                font=(FONT_FAMILY, 9),
                                fill=COLORS["text_secondary"])

        # 범례
        legend = tk.Frame(self._sales_chart_frame, bg=COLORS["card_bg"])
        legend.pack(side=tk.BOTTOM, pady=(0, 4))
        for color, label in zip(COLORS_3, LABELS_3):
            dot = tk.Frame(legend, bg=color, width=12, height=12)
            dot.pack(side=tk.LEFT, padx=(8, 3))
            tk.Label(legend, text=label, bg=COLORS["card_bg"],
                     fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["tiny"])).pack(side=tk.LEFT, padx=(0, 6))

    # ══════════════════════════════════════════════════════════════════════════
    # 탭 2 — BOM 변경이력
    # ══════════════════════════════════════════════════════════════════════════
    def _build_bom_log_tab(self, parent):
        # 필터 바
        fbar = tk.Frame(parent, bg=COLORS["bg"])
        fbar.pack(fill=tk.X, padx=10, pady=8)

        tk.Label(fbar, text="기간:", bg=COLORS["bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._bom_start = tk.Entry(fbar, width=12,
                                   font=(FONT_FAMILY, FONT_SIZES["small"]))
        self._bom_start.insert(0, (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%d"))
        self._bom_start.pack(side=tk.LEFT, padx=(3, 2))
        tk.Label(fbar, text="~", bg=COLORS["bg"]).pack(side=tk.LEFT)
        self._bom_end = tk.Entry(fbar, width=12,
                                 font=(FONT_FAMILY, FONT_SIZES["small"]))
        self._bom_end.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self._bom_end.pack(side=tk.LEFT, padx=(2, 8))

        tk.Label(fbar, text="제품코드:", bg=COLORS["bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._bom_prod_var = tk.StringVar()
        tk.Entry(fbar, textvariable=self._bom_prod_var, width=14,
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT, padx=(3, 8))

        tk.Button(fbar, text="🔍 조회", font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg=COLORS["primary"], fg="white", padx=10, pady=3,
                  cursor="hand2", relief="flat",
                  command=self._reload_bom_log).pack(side=tk.LEFT)
        tk.Button(fbar, text="전체 조회", font=(FONT_FAMILY, FONT_SIZES["small"]),
                  padx=8, pady=3, cursor="hand2",
                  command=lambda: self._reload_bom_log(all_data=True)).pack(side=tk.LEFT, padx=4)

        self._bom_count_lbl = tk.Label(fbar, text="", bg=COLORS["bg"],
                                       fg=COLORS["text_secondary"],
                                       font=(FONT_FAMILY, FONT_SIZES["small"]))
        self._bom_count_lbl.pack(side=tk.LEFT, padx=8)

        # 테이블
        card = tk.Frame(parent, bg=COLORS["card_bg"],
                        highlightbackground=COLORS["border"], highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        cols = ("No", "변경일시", "변경유형", "제품코드", "제품명",
                "부품품번", "부품명", "이전소요량", "변경소요량", "변경자")
        self._bom_tree = ttk.Treeview(card, columns=cols,
                                      show="headings", height=22)
        widths = {"No": 45, "변경일시": 145, "변경유형": 85,
                  "제품코드": 95, "제품명": 150,
                  "부품품번": 110, "부품명": 150,
                  "이전소요량": 90, "변경소요량": 90, "변경자": 80}
        for c in cols:
            self._bom_tree.heading(c, text=c)
            self._bom_tree.column(c, width=widths.get(c, 100), anchor="center")
        self._bom_tree.column("제품명", anchor="w")
        self._bom_tree.column("부품명", anchor="w")

        self._bom_tree.tag_configure("추가", foreground="#16a34a", background="#f0fdf4")
        self._bom_tree.tag_configure("수정", foreground="#1d4ed8", background="#eff6ff")
        self._bom_tree.tag_configure("삭제", foreground="#dc2626", background="#fef2f2")

        sb = ttk.Scrollbar(card, orient="vertical", command=self._bom_tree.yview)
        self._bom_tree.configure(yscrollcommand=sb.set)
        self._bom_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0), pady=5)
        sb.pack(side=tk.RIGHT, fill=tk.Y, pady=5, padx=(0, 5))

        self._reload_bom_log()

    def _reload_bom_log(self, all_data=False):
        if all_data:
            start, end, pid = None, None, None
        else:
            start = self._bom_start.get().strip() or None
            end   = self._bom_end.get().strip()   or None
            pid   = self._bom_prod_var.get().strip() or None

        def load():
            try:
                data = self.app.db.get_bom_change_history(start, end, pid)
                self.app.root.after(0, lambda: self._render_bom_log(data))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err))

        threading.Thread(target=load, daemon=True).start()

    def _render_bom_log(self, data):
        self._bom_log_data = data
        self._bom_tree.delete(*self._bom_tree.get_children())
        for i, r in enumerate(data, 1):
            action = str(r.get("변경유형", ""))
            tag    = action.split("(")[0] if action else "수정"
            old_q  = r.get("이전소요량", "")
            new_q  = r.get("변경소요량", "")
            self._bom_tree.insert("", "end", values=(
                i,
                r.get("변경일시", ""),
                action,
                r.get("제품코드", ""),
                r.get("제품명", ""),
                r.get("부품품번", ""),
                r.get("부품명", ""),
                old_q if old_q != "" else "—",
                new_q if new_q != "" else "—",
                r.get("변경자", ""),
            ), tags=(tag,))
        self._bom_count_lbl.config(text=f"총 {len(data):,}건")

    # ══════════════════════════════════════════════════════════════════════════
    # 탭 3 — 단가 변경이력 (PriceHistoryPage 동일 구현)
    # ══════════════════════════════════════════════════════════════════════════
    def _build_price_log_tab(self, parent):
        # ── 서브 Notebook (부품 구매단가 / 제품 판매단가 탭) ─────────────
        style = ttk.Style()
        style.configure("PH.TNotebook.Tab",
                        font=(FONT_FAMILY, FONT_SIZES["small"], "bold"),
                        padding=[14, 6])
        sub_nb = ttk.Notebook(parent, style="PH.TNotebook")
        sub_nb.pack(fill=tk.BOTH, expand=True)

        part_frame = tk.Frame(sub_nb, bg=COLORS["bg"])
        sub_nb.add(part_frame, text="  🔩 부품 구매단가  ")

        prod_frame = tk.Frame(sub_nb, bg=COLORS["bg"])
        sub_nb.add(prod_frame, text="  📦 제품 판매단가  ")

        self._ph_prod_loaded = False

        def _on_ph_sub_tab(event):
            idx = sub_nb.index(sub_nb.select())
            if idx == 1 and not self._ph_prod_loaded:
                self._ph_prod_loaded = True
                self._build_prod_price_sub_tab(prod_frame)
                self._php_load_data()

        sub_nb.bind("<<NotebookTabChanged>>", _on_ph_sub_tab)
        self._build_part_price_sub_tab(part_frame)

    def _build_part_price_sub_tab(self, parent):
        self._ph_all_records = []
        self._ph_suppliers   = []
        self._ph_sort_col    = "dt"
        self._ph_sort_rev    = True

        # ── 새로고침 버튼 ──────────────────────────────────────────────────
        title_f = tk.Frame(parent, bg=COLORS["bg"])
        title_f.pack(fill=tk.X, padx=10, pady=(8, 4))
        tk.Button(title_f, text="🔄 새로고침",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg="#e2e8f0", fg=COLORS["text"], padx=10, pady=3,
                  cursor="hand2",
                  command=self._ph_load_data).pack(side=tk.RIGHT)

        # ── 대시보드 카드 영역 ─────────────────────────────────────────────
        self._ph_dashboard_frame = tk.Frame(parent, bg=COLORS["bg"])
        self._ph_dashboard_frame.pack(fill=tk.X, padx=10, pady=(0, 6))

        # ── 바 차트 영역 ───────────────────────────────────────────────────
        self._ph_chart_card = tk.Frame(parent, bg=COLORS["card_bg"],
                                       highlightbackground=COLORS["border"],
                                       highlightthickness=1)
        self._ph_chart_card.pack(fill=tk.X, padx=10, pady=(0, 6))

        # ── 필터 바 ────────────────────────────────────────────────────────
        filter_card = tk.Frame(parent, bg=COLORS["card_bg"],
                               highlightbackground=COLORS["border"], highlightthickness=1)
        filter_card.pack(fill=tk.X, padx=10, pady=(0, 4))
        frow = tk.Frame(filter_card, bg=COLORS["card_bg"])
        frow.pack(fill=tk.X, padx=15, pady=10)

        now = datetime.now()
        tk.Label(frow, text="연도:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._ph_year_var = tk.StringVar(value=str(now.year))
        years = [str(y) for y in range(now.year, now.year - 5, -1)]
        ttk.Combobox(frow, textvariable=self._ph_year_var,
                     values=years, width=7, state="readonly").pack(side=tk.LEFT, padx=(4, 12))

        tk.Label(frow, text="월:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._ph_month_var = tk.StringVar(value="전체")
        months = ["전체"] + [str(m) for m in range(1, 13)]
        ttk.Combobox(frow, textvariable=self._ph_month_var,
                     values=months, width=6, state="readonly").pack(side=tk.LEFT, padx=(4, 12))

        tk.Label(frow, text="업체:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._ph_supplier_var = tk.StringVar(value="전체")
        self._ph_supplier_cb = ttk.Combobox(frow, textvariable=self._ph_supplier_var,
                                             values=["전체"], width=14, state="readonly")
        self._ph_supplier_cb.pack(side=tk.LEFT, padx=(4, 12))

        tk.Label(frow, text="구분:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._ph_dir_var = tk.StringVar(value="전체")
        ttk.Combobox(frow, textvariable=self._ph_dir_var,
                     values=["전체", "인상(+)", "인하(-)", "신규"],
                     width=9, state="readonly").pack(side=tk.LEFT, padx=(4, 16))

        tk.Button(frow, text="🔍 조회",
                  font=(FONT_FAMILY, FONT_SIZES["small"], "bold"),
                  bg=COLORS["primary"], fg="white", padx=14, pady=3,
                  cursor="hand2",
                  command=self._ph_apply_filter).pack(side=tk.LEFT, padx=(0, 6))

        tk.Button(frow, text="📥 엑셀 내보내기",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg=COLORS["success"], fg="white", padx=12, pady=3,
                  cursor="hand2",
                  command=self._ph_export_excel).pack(side=tk.LEFT, padx=(0, 6))

        tk.Button(frow, text="📊 월간 보고서",
                  font=(FONT_FAMILY, FONT_SIZES["small"], "bold"),
                  bg=COLORS["info"], fg="white", padx=12, pady=3,
                  cursor="hand2",
                  command=self._ph_export_monthly_report).pack(side=tk.LEFT)

        # ── 테이블 ─────────────────────────────────────────────────────────
        tbl_card = tk.Frame(parent, bg=COLORS["card_bg"],
                            highlightbackground=COLORS["border"], highlightthickness=1)
        tbl_card.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        hdr = tk.Frame(tbl_card, bg=COLORS["card_bg"])
        hdr.pack(fill=tk.X, padx=15, pady=(10, 4))
        tk.Label(hdr, text="상세 변경 이력",
                 bg=COLORS["card_bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")).pack(side=tk.LEFT)
        self._ph_count_label = tk.Label(hdr, text="(0건)",
                                        bg=COLORS["card_bg"], fg=COLORS["text_secondary"],
                                        font=(FONT_FAMILY, FONT_SIZES["small"]))
        self._ph_count_label.pack(side=tk.LEFT, padx=8)

        tf = tk.Frame(tbl_card, bg=COLORS["card_bg"])
        tf.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 12))

        cols = ("no", "dt", "part_id", "part_name", "supplier",
                "old_price", "new_price", "rate", "changed_by", "reason")
        self._ph_tree = ttk.Treeview(tf, columns=cols, show="headings", height=16)
        cfg = [
            ("no",         "No",     40,  "center"),
            ("dt",         "변경일시", 140, "center"),
            ("part_id",    "품번",    100, "center"),
            ("part_name",  "부품명",  160, "w"),
            ("supplier",   "업체명",   90, "w"),
            ("old_price",  "이전단가",  90, "center"),
            ("new_price",  "변경단가",  90, "center"),
            ("rate",       "변경률",   70, "center"),
            ("changed_by", "변경자",   70, "center"),
            ("reason",     "변경사유", 130, "w"),
        ]
        for cid, heading, width, anchor in cfg:
            self._ph_tree.heading(cid, text=heading,
                                  command=lambda c=cid: self._ph_sort_by(c))
            self._ph_tree.column(cid, width=width, anchor=anchor,
                                 minwidth=40, stretch=(cid not in ("no",)))

        vsb = ttk.Scrollbar(tf, orient="vertical", command=self._ph_tree.yview)
        self._ph_tree.configure(yscrollcommand=vsb.set)
        self._ph_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        self._ph_tree.tag_configure("up",   background="#fef2f2", foreground="#dc2626")
        self._ph_tree.tag_configure("down", background="#f0fdf4", foreground="#16a34a")
        self._ph_tree.tag_configure("new",  background="#eff6ff", foreground="#2563eb")

        # 우클릭 삭제 메뉴
        self._ph_row_data = {}   # tree iid → 원본 레코드 dict
        self._ph_menu = tk.Menu(self.app.root, tearoff=0)
        self._ph_menu.add_command(label="🗑 이 이력 삭제", command=self._ph_delete_selected)
        self._ph_tree.bind("<Button-3>", self._ph_right_click)

        self._ph_load_data()

    def _ph_load_data(self):
        def load():
            try:
                records = self.app.db.get_price_history()
                monthly = self.app.db.get_price_history_monthly_summary()
                self.app.root.after(0, lambda: self._ph_on_data_loaded(records, monthly))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err))
        threading.Thread(target=load, daemon=True).start()

    def _ph_on_data_loaded(self, records, monthly):
        self._ph_all_records = records
        suppliers = sorted({r["업체명"] for r in records if r["업체명"]})
        self._ph_suppliers = suppliers
        self._ph_supplier_cb.configure(values=["전체"] + suppliers)
        self._ph_render_dashboard(records, monthly)
        self._ph_render_chart(monthly)
        self._ph_apply_filter()

    def _ph_render_dashboard(self, records, monthly):
        for w in self._ph_dashboard_frame.winfo_children():
            w.destroy()
        now_ym     = datetime.now().strftime("%Y-%m")
        this_month = [r for r in records if r["변경일시"][:7] == now_ym]
        total_m    = len(this_month)
        up_m       = sum(1 for r in this_month if r["변경률"].startswith("+"))
        down_m     = sum(1 for r in this_month if r["변경률"].startswith("-"))
        rates = []
        for r in this_month:
            rate_str = r["변경률"].replace("%", "").replace("+", "")
            try:
                rates.append(float(rate_str))
            except ValueError:
                pass
        avg_rate = sum(rates) / len(rates) if rates else 0
        avg_str  = f"{avg_rate:+.1f}%" if rates else "-"
        stats = [
            ("이번달 변경",  f"{total_m}건",  COLORS["primary"]),
            ("평균 변동률",  avg_str,          COLORS["warning"] if avg_rate >= 0 else COLORS["success"]),
            ("인상 건수",    f"{up_m}건 🔴",   COLORS["danger"]),
            ("인하 건수",    f"{down_m}건 🟢", COLORS["success"]),
        ]
        for label, value, color in stats:
            c = tk.Frame(self._ph_dashboard_frame, bg="white",
                         highlightbackground=color, highlightthickness=2)
            c.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=3)
            tk.Label(c, text=label, bg="white", fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["small"])).pack(pady=(10, 2))
            tk.Label(c, text=value, bg="white", fg=color,
                     font=(FONT_FAMILY, FONT_SIZES["stat"], "bold")).pack(pady=(0, 10))

    def _ph_render_chart(self, monthly):
        for w in self._ph_chart_card.winfo_children():
            w.destroy()
        hdr = tk.Frame(self._ph_chart_card, bg=COLORS["card_bg"])
        hdr.pack(fill=tk.X, padx=15, pady=(10, 4))
        tk.Label(hdr, text="월별 단가 변경 현황 (최근 6개월)",
                 bg=COLORS["card_bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")).pack(side=tk.LEFT)
        now = datetime.now()
        months_6 = []
        for i in range(5, -1, -1):
            m = now.month - i
            y = now.year
            while m <= 0:
                m += 12
                y -= 1
            months_6.append(f"{y}-{str(m).zfill(2)}")
        chart_frame = tk.Frame(self._ph_chart_card, bg=COLORS["card_bg"])
        chart_frame.pack(fill=tk.X, padx=15, pady=(0, 14))
        max_val = max((monthly.get(ym, {}).get("total", 0) for ym in months_6), default=1) or 1
        BAR_W     = 60
        BAR_MAX_H = 100
        CANVAS_H  = BAR_MAX_H + 50
        for ym in months_6:
            data   = monthly.get(ym, {"total": 0, "up": 0, "down": 0})
            total  = data["total"]
            up_cnt = data["up"]
            dn_cnt = data["down"]
            col_f  = tk.Frame(chart_frame, bg=COLORS["card_bg"])
            col_f.pack(side=tk.LEFT, expand=True)
            canvas = tk.Canvas(col_f, width=BAR_W, height=CANVAS_H,
                               bg=COLORS["card_bg"], highlightthickness=0)
            canvas.pack()
            bar_h = int(BAR_MAX_H * total / max_val) if total > 0 else 2
            bot_y = CANVAS_H - 30
            if total > 0:
                up_h  = int(bar_h * up_cnt / total)
                dn_h  = int(bar_h * dn_cnt / total)
                etc_h = bar_h - up_h - dn_h
                x0, x1 = BAR_W // 2 - 14, BAR_W // 2 + 14
                y_cur = bot_y
                for h, color in [(dn_h, "#22c55e"), (etc_h, "#6366f1"), (up_h, "#ef4444")]:
                    if h > 0:
                        canvas.create_rectangle(x0, y_cur - h, x1, y_cur,
                                                fill=color, outline="")
                        y_cur -= h
                canvas.create_text(BAR_W // 2, bot_y - bar_h - 6,
                                   text=str(total), font=(FONT_FAMILY, 9, "bold"),
                                   fill=COLORS["text"])
            else:
                canvas.create_rectangle(
                    BAR_W // 2 - 14, bot_y - 2, BAR_W // 2 + 14, bot_y,
                    fill=COLORS["border"], outline="")
            canvas.create_text(BAR_W // 2, CANVAS_H - 14,
                               text=ym[5:] + "월", font=(FONT_FAMILY, 9),
                               fill=COLORS["text_secondary"])
        legend = tk.Frame(self._ph_chart_card, bg=COLORS["card_bg"])
        legend.pack(pady=(0, 10))
        for color, text in [("#ef4444", "인상"), ("#22c55e", "인하"), ("#6366f1", "신규")]:
            tk.Frame(legend, bg=color, width=12, height=12).pack(side=tk.LEFT, padx=(8, 2))
            tk.Label(legend, text=text, bg=COLORS["card_bg"],
                     fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT, padx=(0, 6))

    def _ph_apply_filter(self):
        year     = self._ph_year_var.get()
        month    = self._ph_month_var.get()
        supplier = self._ph_supplier_var.get()
        dir_sel  = self._ph_dir_var.get()
        filtered = []
        for r in self._ph_all_records:
            dt = r["변경일시"]
            if not dt.startswith(year):
                continue
            if month != "전체":
                ym = f"{year}-{month.zfill(2)}"
                if not dt.startswith(ym):
                    continue
            if supplier != "전체" and r["업체명"] != supplier:
                continue
            rate = r["변경률"]
            if dir_sel == "인상(+)" and not rate.startswith("+"):
                continue
            if dir_sel == "인하(-)" and not rate.startswith("-"):
                continue
            if dir_sel == "신규" and rate != "신규":
                continue
            filtered.append(r)
        filtered.sort(key=lambda x: x.get("변경일시", ""), reverse=True)
        self._ph_render_table(filtered)

    def _ph_render_table(self, rows):
        self._ph_tree.delete(*self._ph_tree.get_children())
        self._ph_row_data = {}
        self._ph_count_label.configure(text=f"({len(rows)}건)")
        for i, r in enumerate(rows, 1):
            rate = r["변경률"]
            tag  = "up" if rate.startswith("+") else ("down" if rate.startswith("-") else "new")
            old_p = f"{r['이전단가']:,.0f}" if r['이전단가'] else "-"
            new_p = f"{r['변경단가']:,.0f}" if r['변경단가'] else "-"
            iid = self._ph_tree.insert("", tk.END, values=(
                i, r["변경일시"], r["품번"], r["부품명"], r["업체명"],
                old_p, new_p, rate, r["변경자"], r["변경사유"],
            ), tags=(tag,))
            self._ph_row_data[iid] = r

    def _ph_right_click(self, event):
        iid = self._ph_tree.identify_row(event.y)
        if not iid:
            return
        self._ph_tree.selection_set(iid)
        self._ph_menu.post(event.x_root, event.y_root)

    def _ph_delete_selected(self):
        selected = self._ph_tree.selection()
        if not selected:
            return
        iid = selected[0]
        r = self._ph_row_data.get(iid)
        if not r:
            return

        old_p = f"{r['이전단가']:,.0f}" if r['이전단가'] else "-"
        new_p = f"{r['변경단가']:,.0f}" if r['변경단가'] else "-"
        info = (f"변경일시: {r['변경일시']}\n"
                f"품번: {r['품번']} ({r['부품명']})\n"
                f"업체명: {r['업체명']}\n"
                f"{old_p} → {new_p}원 ({r['변경률']})")
        if not messagebox.askyesno("이력 삭제",
                                    f"다음 단가변경이력을 삭제하시겠습니까?\n\n{info}\n\n"
                                    "※ 이 작업은 이력 기록만 제거하며 부품 현재 단가에는 영향을 주지 않습니다."):
            return

        def work():
            try:
                ok = self.app.db.delete_price_history(
                    r["변경일시"], r["품번"], r["이전단가"], r["변경단가"])
                self.app.root.after(0, lambda: self._ph_delete_done(ok))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err))

        threading.Thread(target=work, daemon=True).start()

    def _ph_delete_done(self, ok):
        if ok:
            messagebox.showinfo("삭제 완료", "이력이 삭제되었습니다.")
            self._ph_load_data()
        else:
            messagebox.showerror("삭제 실패", "해당 이력을 찾을 수 없습니다. 목록을 새로고침 후 다시 시도해주세요.")

    def _ph_sort_by(self, col):
        if self._ph_sort_col == col:
            self._ph_sort_rev = not self._ph_sort_rev
        else:
            self._ph_sort_col = col
            self._ph_sort_rev = True
        self._ph_apply_filter()

    def _ph_export_excel(self):
        rows = [self._ph_tree.item(iid)["values"]
                for iid in self._ph_tree.get_children()]
        if not rows:
            messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
            return
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            title="단가변경이력 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=f"단가변경이력_{now_str}.xlsx",
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "단가변경이력"
            hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            hdr_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
            nrm_font = Font(name="맑은 고딕", size=10)
            up_fill  = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
            dn_fill  = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
            nw_fill  = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            thin     = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"),  bottom=Side(style="thin"))
            headers = ["No", "변경일시", "품번", "부품명", "업체명",
                       "이전단가", "변경단가", "변경률", "변경자", "변경사유"]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin
            for r_idx, row in enumerate(rows, 2):
                rate_val = str(row[7])
                fill = up_fill if rate_val.startswith("+") else (
                       dn_fill if rate_val.startswith("-") else nw_fill)
                for c_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.font  = nrm_font
                    cell.fill  = fill
                    cell.border = thin
                    cell.alignment = Alignment(horizontal="center")
            widths = [6, 18, 14, 22, 14, 12, 12, 10, 10, 20]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            wb.save(path)
            messagebox.showinfo("저장 완료", f"저장되었습니다.\n{path}")
            import os
            os.startfile(path)
        except Exception as e:
            messagebox.showerror("오류", str(e))

    # ══════════════════════════════════════════════════════════════════════════
    # 월간 단가변경 보고서 (엑셀, 부품별 추이 그래프 + 제품 원가영향 포함)
    # ══════════════════════════════════════════════════════════════════════════
    def _ph_export_monthly_report(self):
        yr_str = self._ph_year_var.get()
        mo_str = self._ph_month_var.get()
        if mo_str == "전체":
            messagebox.showinfo("알림", "월간 보고서는 특정 월을 선택한 뒤 생성해주세요.\n"
                                        "(현재 '월' 필터가 '전체'로 설정되어 있습니다)")
            return
        year, month = int(yr_str), int(mo_str)

        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            title="월간 단가변경 보고서 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=f"단가변경_월간보고서_{year}{month:02d}_{now_str}.xlsx",
        )
        if not path:
            return

        def work():
            try:
                data = self.app.db.get_monthly_price_change_report_data(year, month)
                # 부품별 전체 이력(추이 그래프용)은 이미 로드된 전체 이력에서 클라이언트 필터링
                full_history = self._ph_all_records or self.app.db.get_price_history()
                self._ph_build_report_workbook(path, year, month, data, full_history)
                self.app.root.after(0, lambda: self._ph_report_done(path))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err))

        threading.Thread(target=work, daemon=True).start()

    def _ph_report_done(self, path):
        messagebox.showinfo("생성 완료", f"월간 보고서가 저장되었습니다.\n{path}")
        import os
        os.startfile(path)

    def _ph_build_report_workbook(self, path, year, month, data, full_history):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import LineChart, BarChart, Reference
        from openpyxl.chart.marker import Marker
        from openpyxl.utils import get_column_letter

        ym = f"{year}-{month:02d}"
        changed_parts    = data["changed_parts"]
        product_impacts  = data["product_impacts"]
        raw_rows         = data["raw_rows"]

        TITLE_FONT = Font(name="맑은 고딕", bold=True, size=16, color="1E293B")
        SUB_FONT   = Font(name="맑은 고딕", size=10, color="64748B")
        HDR_FILL   = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        HDR_FONT   = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
        NRM_FONT   = Font(name="맑은 고딕", size=10)
        BOLD_FONT  = Font(name="맑은 고딕", bold=True, size=10)
        UP_FILL    = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
        DN_FILL    = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
        CARD_FILL  = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        THIN       = Border(left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin"))
        CENTER     = Alignment(horizontal="center", vertical="center")

        def style_header_row(ws, row, ncols):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = HDR_FONT
                cell.fill = HDR_FILL
                cell.alignment = CENTER
                cell.border = THIN

        wb = openpyxl.Workbook()

        # ────────────────────────────────────────────────────────────────
        # 시트 1: 월간 요약
        # ────────────────────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "월간 요약"

        ws1["B2"] = f"{ym} 단가변경 월간 보고서"
        ws1["B2"].font = TITLE_FONT
        ws1["B3"] = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws1["B3"].font = SUB_FONT

        total_cnt = len(raw_rows)
        up_cnt    = sum(1 for r in raw_rows if r["변경률"].startswith("+"))
        dn_cnt    = sum(1 for r in raw_rows if r["변경률"].startswith("-"))
        new_cnt   = sum(1 for r in raw_rows if r["변경률"] == "신규")
        part_cnt  = len(changed_parts)
        total_impact = sum(p["증감액"] for p in product_impacts)

        stat_row = 5
        stats = [
            ("변경 부품 수", f"{part_cnt}종"),
            ("총 변경 건수", f"{total_cnt}건"),
            ("인상 / 인하 / 신규", f"{up_cnt} / {dn_cnt} / {new_cnt}"),
            ("영향받은 제품 수", f"{len(product_impacts)}종"),
            ("제품 원가 증감 합계", f"{total_impact:+,}원"),
        ]
        for i, (label, val) in enumerate(stats):
            r = stat_row + i
            ws1.cell(row=r, column=2, value=label).font = BOLD_FONT
            ws1.cell(row=r, column=2).fill = CARD_FILL
            ws1.cell(row=r, column=2).border = THIN
            ws1.cell(row=r, column=3, value=val).font = NRM_FONT
            ws1.cell(row=r, column=3).border = THIN

        # 부품별 변경 요약 테이블
        t1_row = stat_row + len(stats) + 2
        ws1.cell(row=t1_row, column=2, value="■ 부품별 단가변경 요약").font = BOLD_FONT
        hdr_row = t1_row + 1
        headers1 = ["품번", "부품명", "업체명", "이전단가", "최종단가", "변경률", "변경횟수"]
        for c, h in enumerate(headers1, 2):
            ws1.cell(row=hdr_row, column=c, value=h)
        style_header_row_range = lambda r, c0, c1: [
            (lambda cell: (setattr(cell, "font", HDR_FONT),
                           setattr(cell, "fill", HDR_FILL),
                           setattr(cell, "alignment", CENTER),
                           setattr(cell, "border", THIN)))(ws1.cell(row=r, column=c))
            for c in range(c0, c1 + 1)
        ]
        style_header_row_range(hdr_row, 2, 2 + len(headers1) - 1)

        r = hdr_row + 1
        for cp in changed_parts:
            rate = cp["변경률"]
            rate_str = "신규" if rate is None else f"{rate:+.1f}%"
            fill = UP_FILL if (rate is not None and rate > 0) else (
                   DN_FILL if (rate is not None and rate < 0) else CARD_FILL)
            vals = [cp["품번"], cp["부품명"], cp["업체명"],
                    f"{cp['이전단가']:,.0f}", f"{cp['변경단가']:,.0f}",
                    rate_str, cp["변경횟수"]]
            for c, v in enumerate(vals, 2):
                cell = ws1.cell(row=r, column=c, value=v)
                cell.font = NRM_FONT
                cell.fill = fill
                cell.border = THIN
                cell.alignment = CENTER
            r += 1

        # 제품 원가 영향 테이블
        t2_row = r + 2
        ws1.cell(row=t2_row, column=2, value="■ 단가변경에 따른 제품 원가 영향").font = BOLD_FONT
        hdr2_row = t2_row + 1
        headers2 = ["제품코드", "제품명", "원가(변경전)", "원가(변경후)", "증감액", "증감률", "영향부품"]
        for c, h in enumerate(headers2, 2):
            ws1.cell(row=hdr2_row, column=c, value=h)
        style_header_row_range(hdr2_row, 2, 2 + len(headers2) - 1)

        r = hdr2_row + 1
        impact_table_start = r
        for pi in product_impacts:
            rate = pi["증감률"]
            rate_str = "-" if rate is None else f"{rate:+.1f}%"
            fill = UP_FILL if pi["증감액"] > 0 else DN_FILL
            part_names = ", ".join(a["부품명"] for a in pi["영향부품"])
            vals = [pi["제품코드"], pi["제품명"],
                    f"{pi['원가변경전']:,}", f"{pi['원가변경후']:,}",
                    f"{pi['증감액']:+,}", rate_str, part_names]
            for c, v in enumerate(vals, 2):
                cell = ws1.cell(row=r, column=c, value=v)
                cell.font = NRM_FONT
                cell.fill = fill
                cell.border = THIN
                cell.alignment = CENTER if c != 8 else Alignment(horizontal="left", vertical="center")
            r += 1
        impact_table_end = r - 1

        for col, w in zip("BCDEFGH", [14, 20, 14, 14, 14, 10, 10]):
            ws1.column_dimensions[col].width = w
        ws1.column_dimensions["H"].width = 30

        # 원가 영향 상위 제품 바 차트
        if product_impacts:
            top_n = min(10, len(product_impacts))
            chart_data_row = impact_table_end + 3
            ws1.cell(row=chart_data_row, column=2, value="■ 원가 증감 상위 제품 (Top 10)").font = BOLD_FONT
            data_hdr = chart_data_row + 1
            ws1.cell(row=data_hdr, column=2, value="제품명")
            ws1.cell(row=data_hdr, column=3, value="증감액")
            style_header_row_range(data_hdr, 2, 3)
            for i in range(top_n):
                pi = product_impacts[i]
                ws1.cell(row=data_hdr + 1 + i, column=2, value=pi["제품명"] or pi["제품코드"])
                ws1.cell(row=data_hdr + 1 + i, column=3, value=pi["증감액"])

            bar = BarChart()
            bar.type = "bar"
            bar.title = "제품별 원가 증감액 (Top 10)"
            bar.y_axis.title = "제품"
            bar.x_axis.title = "증감액(원)"
            bar.height = 1.8 + 0.7 * top_n
            bar.width = 18
            cats = Reference(ws1, min_col=2, min_row=data_hdr + 1, max_row=data_hdr + top_n)
            vals = Reference(ws1, min_col=3, min_row=data_hdr, max_row=data_hdr + top_n)
            bar.add_data(vals, titles_from_data=True)
            bar.set_categories(cats)
            bar.legend = None
            ws1.add_chart(bar, f"E{chart_data_row}")

        # ────────────────────────────────────────────────────────────────
        # 시트 2: 부품별 단가 추이 (전체 이력 기준 라인 차트)
        # ────────────────────────────────────────────────────────────────
        ws2 = wb.create_sheet("부품별 단가추이")
        ws2["B2"] = f"부품별 단가 변동 추이 (변경이력 전체, {ym}월 변경 부품 대상)"
        ws2["B2"].font = TITLE_FONT

        cur_row = 4
        for cp in changed_parts:
            pid = cp["품번"]
            hist = sorted(
                (r for r in full_history if r["품번"] == pid),
                key=lambda r: r["변경일시"],
            )
            if not hist:
                continue

            ws2.cell(row=cur_row, column=2,
                    value=f"[{pid}] {cp['부품명']} ({cp['업체명']})").font = BOLD_FONT
            data_hdr = cur_row + 1
            ws2.cell(row=data_hdr, column=2, value="일시")
            ws2.cell(row=data_hdr, column=3, value="단가")
            style_header_row_range2 = lambda r_, c0, c1, sheet=ws2: [
                (lambda cell: (setattr(cell, "font", HDR_FONT),
                               setattr(cell, "fill", HDR_FILL),
                               setattr(cell, "alignment", CENTER),
                               setattr(cell, "border", THIN)))(sheet.cell(row=r_, column=c))
                for c in range(c0, c1 + 1)
            ]
            style_header_row_range2(data_hdr, 2, 3)

            # 최초 이전단가를 시작점으로 포함해 변화 추이를 온전히 표현
            points = [(hist[0]["변경일시"][:10] + " (이전)", hist[0]["이전단가"])]
            for h in hist:
                points.append((h["변경일시"][:10], h["변경단가"]))

            r = data_hdr + 1
            for label, price in points:
                ws2.cell(row=r, column=2, value=label).font = NRM_FONT
                ws2.cell(row=r, column=3, value=price).font = NRM_FONT
                ws2.cell(row=r, column=2).border = THIN
                ws2.cell(row=r, column=3).border = THIN
                r += 1
            data_end = r - 1

            line = LineChart()
            line.title = f"{cp['부품명']} 단가 추이"
            line.style = 12
            line.y_axis.title = "단가(원)"
            line.x_axis.title = "변경일"
            line.height = 6
            line.width = 14
            vals = Reference(ws2, min_col=3, min_row=data_hdr, max_row=data_end)
            cats = Reference(ws2, min_col=2, min_row=data_hdr + 1, max_row=data_end)
            line.add_data(vals, titles_from_data=True)
            line.set_categories(cats)
            series = line.series[0]
            series.marker = Marker(symbol="circle", size=6)
            series.smooth = False
            line.legend = None
            ws2.add_chart(line, f"E{cur_row}")

            block_h = max(data_end - cur_row + 2, 13)
            cur_row += block_h + 1

        ws2.column_dimensions["B"].width = 18
        ws2.column_dimensions["C"].width = 12

        # ────────────────────────────────────────────────────────────────
        # 시트 3: 상세 변경 이력 (원본)
        # ────────────────────────────────────────────────────────────────
        ws3 = wb.create_sheet("상세이력")
        headers3 = ["No", "변경일시", "품번", "부품명", "업체명",
                    "이전단가", "변경단가", "변경률", "변경자", "변경사유"]
        for c, h in enumerate(headers3, 1):
            ws3.cell(row=1, column=c, value=h)
        style_header_row(ws3, 1, len(headers3))
        for i, rrow in enumerate(raw_rows, 1):
            rate = rrow["변경률"]
            fill = UP_FILL if rate.startswith("+") else (DN_FILL if rate.startswith("-") else CARD_FILL)
            vals = [i, rrow["변경일시"], rrow["품번"], rrow["부품명"], rrow["업체명"],
                    f"{rrow['이전단가']:,.0f}", f"{rrow['변경단가']:,.0f}",
                    rate, rrow["변경자"], rrow["변경사유"]]
            for c, v in enumerate(vals, 1):
                cell = ws3.cell(row=i + 1, column=c, value=v)
                cell.font = NRM_FONT
                cell.fill = fill
                cell.border = THIN
                cell.alignment = CENTER
        widths3 = [6, 18, 14, 22, 14, 12, 12, 10, 10, 20]
        for i, w in enumerate(widths3, 1):
            ws3.column_dimensions[get_column_letter(i)].width = w

        wb.save(path)

    # ══════════════════════════════════════════════════════════════════════════
    # 단가변경이력 탭 — 제품 판매단가 서브탭
    # ══════════════════════════════════════════════════════════════════════════
    def _build_prod_price_sub_tab(self, parent):
        self._php_all_records = []

        # ── 새로고침 버튼 ─────────────────────────────────────────────────
        title_f = tk.Frame(parent, bg=COLORS["bg"])
        title_f.pack(fill=tk.X, padx=10, pady=(8, 4))
        tk.Button(title_f, text="🔄 새로고침",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg="#e2e8f0", fg=COLORS["text"], padx=10, pady=3,
                  cursor="hand2",
                  command=self._php_load_data).pack(side=tk.RIGHT)

        # ── 대시보드 카드 ─────────────────────────────────────────────────
        self._php_dashboard_frame = tk.Frame(parent, bg=COLORS["bg"])
        self._php_dashboard_frame.pack(fill=tk.X, padx=10, pady=(0, 6))

        # ── 필터 바 ───────────────────────────────────────────────────────
        filter_card = tk.Frame(parent, bg=COLORS["card_bg"],
                               highlightbackground=COLORS["border"], highlightthickness=1)
        filter_card.pack(fill=tk.X, padx=10, pady=(0, 4))
        frow = tk.Frame(filter_card, bg=COLORS["card_bg"])
        frow.pack(fill=tk.X, padx=15, pady=10)

        now = datetime.now()
        tk.Label(frow, text="연도:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._php_year_var = tk.StringVar(value=str(now.year))
        years = [str(y) for y in range(now.year, now.year - 5, -1)]
        ttk.Combobox(frow, textvariable=self._php_year_var,
                     values=years, width=7, state="readonly").pack(side=tk.LEFT, padx=(4, 12))

        tk.Label(frow, text="월:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._php_month_var = tk.StringVar(value="전체")
        months = ["전체"] + [str(m) for m in range(1, 13)]
        ttk.Combobox(frow, textvariable=self._php_month_var,
                     values=months, width=6, state="readonly").pack(side=tk.LEFT, padx=(4, 12))

        tk.Label(frow, text="구분:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._php_dir_var = tk.StringVar(value="전체")
        ttk.Combobox(frow, textvariable=self._php_dir_var,
                     values=["전체", "인상(+)", "인하(-)", "신규"],
                     width=9, state="readonly").pack(side=tk.LEFT, padx=(4, 16))

        tk.Button(frow, text="🔍 조회",
                  font=(FONT_FAMILY, FONT_SIZES["small"], "bold"),
                  bg=COLORS["primary"], fg="white", padx=14, pady=3,
                  cursor="hand2",
                  command=self._php_apply_filter).pack(side=tk.LEFT, padx=(0, 6))

        tk.Button(frow, text="📥 엑셀 내보내기",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg=COLORS["success"], fg="white", padx=12, pady=3,
                  cursor="hand2",
                  command=self._php_export_excel).pack(side=tk.LEFT)

        # ── 테이블 ────────────────────────────────────────────────────────
        tbl_card = tk.Frame(parent, bg=COLORS["card_bg"],
                            highlightbackground=COLORS["border"], highlightthickness=1)
        tbl_card.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        hdr = tk.Frame(tbl_card, bg=COLORS["card_bg"])
        hdr.pack(fill=tk.X, padx=15, pady=(10, 4))
        tk.Label(hdr, text="제품 판매단가 변경 이력",
                 bg=COLORS["card_bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")).pack(side=tk.LEFT)
        self._php_count_label = tk.Label(hdr, text="(0건)",
                                         bg=COLORS["card_bg"], fg=COLORS["text_secondary"],
                                         font=(FONT_FAMILY, FONT_SIZES["small"]))
        self._php_count_label.pack(side=tk.LEFT, padx=8)

        tf = tk.Frame(tbl_card, bg=COLORS["card_bg"])
        tf.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 12))

        cols = ("no", "dt", "prod_id", "prod_name",
                "old_price", "new_price", "rate", "changed_by", "reason")
        self._php_tree = ttk.Treeview(tf, columns=cols, show="headings", height=16)
        cfg = [
            ("no",         "No",        40,  "center"),
            ("dt",         "변경일시",   140, "center"),
            ("prod_id",    "제품코드",   110, "center"),
            ("prod_name",  "제품명",     200, "w"),
            ("old_price",  "이전판매가", 100, "center"),
            ("new_price",  "변경판매가", 100, "center"),
            ("rate",       "변경률",      70, "center"),
            ("changed_by", "변경자",      70, "center"),
            ("reason",     "변경사유",   130, "w"),
        ]
        for cid, heading, width, anchor in cfg:
            self._php_tree.heading(cid, text=heading)
            self._php_tree.column(cid, width=width, anchor=anchor,
                                  minwidth=40, stretch=(cid not in ("no",)))

        vsb = ttk.Scrollbar(tf, orient="vertical", command=self._php_tree.yview)
        self._php_tree.configure(yscrollcommand=vsb.set)
        self._php_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        self._php_tree.tag_configure("up",   background="#fef2f2", foreground="#dc2626")
        self._php_tree.tag_configure("down", background="#f0fdf4", foreground="#16a34a")
        self._php_tree.tag_configure("new",  background="#eff6ff", foreground="#2563eb")

    def _php_load_data(self):
        def load():
            try:
                records = self.app.db.get_product_price_history()
                monthly = self.app.db.get_product_price_history_monthly_summary()
                self.app.root.after(0, lambda: self._php_on_data_loaded(records, monthly))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err))
        threading.Thread(target=load, daemon=True).start()

    def _php_on_data_loaded(self, records, monthly):
        self._php_all_records = records
        self._php_render_dashboard(records, monthly)
        self._php_apply_filter()

    def _php_render_dashboard(self, records, monthly):
        for w in self._php_dashboard_frame.winfo_children():
            w.destroy()
        now_ym     = datetime.now().strftime("%Y-%m")
        this_month = [r for r in records if r["변경일시"][:7] == now_ym]
        total_m    = len(this_month)
        up_m       = sum(1 for r in this_month if r["변경률"].startswith("+"))
        down_m     = sum(1 for r in this_month if r["변경률"].startswith("-"))
        rates = []
        for r in this_month:
            rate_str = r["변경률"].replace("%", "").replace("+", "")
            try:
                rates.append(float(rate_str))
            except ValueError:
                pass
        avg_rate = sum(rates) / len(rates) if rates else 0
        avg_str  = f"{avg_rate:+.1f}%" if rates else "-"
        stats = [
            ("이번달 변경",  f"{total_m}건",  COLORS["primary"]),
            ("평균 변동률",  avg_str,          COLORS["warning"] if avg_rate >= 0 else COLORS["success"]),
            ("인상 건수",    f"{up_m}건 🔴",   COLORS["danger"]),
            ("인하 건수",    f"{down_m}건 🟢", COLORS["success"]),
        ]
        for label, value, color in stats:
            c = tk.Frame(self._php_dashboard_frame, bg="white",
                         highlightbackground=color, highlightthickness=2)
            c.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=3)
            tk.Label(c, text=label, bg="white", fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["small"])).pack(pady=(10, 2))
            tk.Label(c, text=value, bg="white", fg=color,
                     font=(FONT_FAMILY, FONT_SIZES["stat"], "bold")).pack(pady=(0, 10))

    def _php_apply_filter(self):
        year    = self._php_year_var.get()
        month   = self._php_month_var.get()
        dir_sel = self._php_dir_var.get()
        filtered = []
        for r in self._php_all_records:
            dt = r["변경일시"]
            if not dt.startswith(year):
                continue
            if month != "전체":
                ym = f"{year}-{month.zfill(2)}"
                if not dt.startswith(ym):
                    continue
            rate = r["변경률"]
            if dir_sel == "인상(+)" and not rate.startswith("+"):
                continue
            if dir_sel == "인하(-)" and not rate.startswith("-"):
                continue
            if dir_sel == "신규" and rate != "신규":
                continue
            filtered.append(r)
        filtered.sort(key=lambda x: x.get("변경일시", ""), reverse=True)
        self._php_render_table(filtered)

    def _php_render_table(self, rows):
        self._php_tree.delete(*self._php_tree.get_children())
        self._php_count_label.configure(text=f"({len(rows)}건)")
        for i, r in enumerate(rows, 1):
            rate = r["변경률"]
            tag  = "up" if rate.startswith("+") else ("down" if rate.startswith("-") else "new")
            old_p = f"{r['이전판매가']:,.0f}" if r.get("이전판매가") else "-"
            new_p = f"{r['변경판매가']:,.0f}" if r.get("변경판매가") else "-"
            self._php_tree.insert("", tk.END, values=(
                i, r["변경일시"], r["제품코드"], r["제품명"],
                old_p, new_p, rate, r.get("변경자", ""), r.get("변경사유", ""),
            ), tags=(tag,))

    def _php_export_excel(self):
        rows = [self._php_tree.item(iid)["values"]
                for iid in self._php_tree.get_children()]
        if not rows:
            messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
            return
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            title="제품판매단가 변경이력 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=f"제품판매단가_변경이력_{now_str}.xlsx",
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "제품판매단가이력"
            hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            hdr_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
            nrm_font = Font(name="맑은 고딕", size=10)
            up_fill  = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
            dn_fill  = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
            nw_fill  = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            thin     = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"), bottom=Side(style="thin"))
            headers = ["No", "변경일시", "제품코드", "제품명",
                       "이전판매가", "변경판매가", "변경률", "변경자", "변경사유"]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin
            for r_idx, row in enumerate(rows, 2):
                rate_val = str(row[6])
                fill = up_fill if rate_val.startswith("+") else (
                       dn_fill if rate_val.startswith("-") else nw_fill)
                for c_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.font  = nrm_font
                    cell.fill  = fill
                    cell.border = thin
                    cell.alignment = Alignment(horizontal="center")
            widths = [6, 18, 14, 26, 12, 12, 10, 10, 24]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            wb.save(path)
            messagebox.showinfo("저장 완료", f"저장되었습니다.\n{path}")
            import os
            os.startfile(path)
        except Exception as e:
            messagebox.showerror("오류", str(e))

    # ══════════════════════════════════════════════════════════════════════════
    # 탭 4 — 부품 입출고 요약
    # ══════════════════════════════════════════════════════════════════════════
    def _build_io_tab(self, parent):
        # 필터 바
        fbar = tk.Frame(parent, bg=COLORS["bg"])
        fbar.pack(fill=tk.X, padx=10, pady=8)

        tk.Label(fbar, text="조회 연도:", bg=COLORS["bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        years = [str(y) for y in range(datetime.now().year, datetime.now().year - 5, -1)]
        self._io_year_var = tk.StringVar(value=str(self._io_year))
        ttk.Combobox(fbar, textvariable=self._io_year_var,
                     values=years, width=8, state="readonly").pack(side=tk.LEFT, padx=(3, 10))

        tk.Button(fbar, text="🔄 조회", font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg=COLORS["primary"], fg="white", padx=10, pady=3,
                  cursor="hand2", relief="flat",
                  command=self._reload_io).pack(side=tk.LEFT)

        # 요약 카드
        self._io_card_frame = tk.Frame(parent, bg=COLORS["bg"])
        self._io_card_frame.pack(fill=tk.X, padx=10, pady=(0, 8))

        # 차트 영역
        chart_outer = tk.Frame(parent, bg=COLORS["card_bg"],
                               highlightbackground=COLORS["border"],
                               highlightthickness=1)
        chart_outer.pack(fill=tk.X, padx=10, pady=(0, 8))
        tk.Label(chart_outer, text="월별 부품 입출고 현황",
                 bg=COLORS["card_bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")
                 ).pack(anchor="w", padx=15, pady=(10, 4))
        self._io_chart_frame = tk.Frame(chart_outer, bg=COLORS["card_bg"])
        self._io_chart_frame.pack(fill=tk.X, padx=10, pady=(0, 10))

        # 테이블
        tbl_outer = tk.Frame(parent, bg=COLORS["card_bg"],
                             highlightbackground=COLORS["border"],
                             highlightthickness=1)
        tbl_outer.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        cols = ("월", "입고 건수", "입고 수량", "출고 건수", "출고 수량", "순입고 수량")
        self._io_tree = ttk.Treeview(tbl_outer, columns=cols,
                                     show="headings", height=14)
        for c in cols:
            self._io_tree.heading(c, text=c)
            self._io_tree.column(c, width=110, anchor="center")
        self._io_tree.column("월", width=85)

        self._io_tree.tag_configure("in",  foreground="#1d4ed8")
        self._io_tree.tag_configure("out", foreground="#dc2626")
        self._io_tree.tag_configure("sum", background="#f1f5f9",
                                    font=(FONT_FAMILY, FONT_SIZES["small"], "bold"))

        sb = ttk.Scrollbar(tbl_outer, orient="vertical", command=self._io_tree.yview)
        self._io_tree.configure(yscrollcommand=sb.set)
        self._io_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0), pady=5)
        sb.pack(side=tk.RIGHT, fill=tk.Y, pady=5, padx=(0, 5))

        self._reload_io()

    def _reload_io(self):
        try:
            self._io_year = int(self._io_year_var.get())
        except Exception:
            pass

        def load():
            try:
                data = self.app.db.get_parts_io_summary(self._io_year)
                self.app.root.after(0, lambda: self._render_io(data))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err))

        threading.Thread(target=load, daemon=True).start()

    def _render_io(self, data):
        self._io_data = data

        # ── 요약 카드 ────────────────────────────────────────────────────────
        for w in self._io_card_frame.winfo_children():
            w.destroy()

        total_in_cnt  = sum(d["입고건수"] for d in data)
        total_out_cnt = sum(d["출고건수"] for d in data)
        total_in_qty  = sum(d["입고수량"] for d in data)
        total_out_qty = sum(d["출고수량"] for d in data)

        cards = [
            ("연간 입고 건수", f"{total_in_cnt:,}건",  COLORS["primary"]),
            ("연간 출고 건수", f"{total_out_cnt:,}건", COLORS["warning"]),
            ("연간 입고 수량", f"{total_in_qty:,}개",  COLORS["info"]),
            ("연간 출고 수량", f"{total_out_qty:,}개", COLORS["danger"]),
        ]
        for label, val, color in cards:
            card = tk.Frame(self._io_card_frame, bg="white",
                            highlightbackground=color, highlightthickness=2)
            card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=3)
            tk.Label(card, text=label, bg="white",
                     fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["small"])).pack(pady=(8, 2))
            tk.Label(card, text=val, bg="white", fg=color,
                     font=(FONT_FAMILY, FONT_SIZES["stat"], "bold")).pack(pady=(0, 8))

        # ── 차트 ─────────────────────────────────────────────────────────────
        for w in self._io_chart_frame.winfo_children():
            w.destroy()
        self._draw_io_chart(data)

        # ── 테이블 ───────────────────────────────────────────────────────────
        self._io_tree.delete(*self._io_tree.get_children())
        for d in reversed(data):
            net = d["입고수량"] - d["출고수량"]
            self._io_tree.insert("", "end", values=(
                d["월"],
                f"{d['입고건수']:,}",
                f"{d['입고수량']:,}",
                f"{d['출고건수']:,}",
                f"{d['출고수량']:,}",
                f"{net:+,}",
            ))

        # 합계
        if data:
            net_total = total_in_qty - total_out_qty
            self._io_tree.insert("", "end", values=(
                "합계",
                f"{total_in_cnt:,}",
                f"{total_in_qty:,}",
                f"{total_out_cnt:,}",
                f"{total_out_qty:,}",
                f"{net_total:+,}",
            ), tags=("sum",))

    def _draw_io_chart(self, data):
        """입고(파랑)/출고(주황) 그룹 바 차트"""
        if not data:
            tk.Label(self._io_chart_frame, text="데이터 없음",
                     bg=COLORS["card_bg"], fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["small"])).pack(pady=30)
            return

        COL_W   = 80
        BAR_MAX = 100
        CH      = BAR_MAX + 55

        max_val = max(
            max((d["입고수량"] for d in data), default=0),
            max((d["출고수량"] for d in data), default=0),
            1,
        )

        for d in data:
            col_f = tk.Frame(self._io_chart_frame, bg=COLORS["card_bg"])
            col_f.pack(side=tk.LEFT, expand=True)

            canvas = tk.Canvas(col_f, width=COL_W, height=CH,
                               bg=COLORS["card_bg"], highlightthickness=0)
            canvas.pack()

            bar_w = 16
            gap   = 4
            total_bw = 2 * bar_w + gap
            x0    = (COL_W - total_bw) // 2
            bot_y = CH - 30

            for bi, (qty, color) in enumerate([
                (d["입고수량"], "#3b82f6"),
                (d["출고수량"], "#f59e0b"),
            ]):
                bx0 = x0 + bi * (bar_w + gap)
                bx1 = bx0 + bar_w
                h   = max(2, int(BAR_MAX * qty / max_val)) if qty > 0 else 2
                canvas.create_rectangle(bx0, bot_y - h, bx1, bot_y,
                                        fill=color, outline="")

            canvas.create_text(COL_W // 2, CH - 14,
                                text=d["월"][5:] + "월",
                                font=(FONT_FAMILY, 9),
                                fill=COLORS["text_secondary"])

        legend = tk.Frame(self._io_chart_frame, bg=COLORS["card_bg"])
        legend.pack(side=tk.BOTTOM, pady=(0, 4))
        for color, lbl in [("#3b82f6", "입고"), ("#f59e0b", "출고")]:
            dot = tk.Frame(legend, bg=color, width=12, height=12)
            dot.pack(side=tk.LEFT, padx=(8, 3))
            tk.Label(legend, text=lbl, bg=COLORS["card_bg"],
                     fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["tiny"])).pack(side=tk.LEFT, padx=(0, 6))
