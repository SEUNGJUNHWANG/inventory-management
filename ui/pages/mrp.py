"""
재고관리 시스템 - MRP(자재소요계획) 페이지
- 생산 계획 입력 (자동완성 검색)
- 소요 부품 계산 및 발주 리스트 출력
- 안전재고 반영 옵션
- 최대 생산 가능 수량 표시
- 엑셀 내보내기
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import os
from datetime import datetime
from core.constants import COLORS, FONT_FAMILY, FONT_SIZES
from ui.widget_utils import flash_btn


class MrpPage:
    """MRP(자재소요계획) 페이지"""

    def __init__(self, app):
        self.app = app
        self.production_plan = []  # [{"product_id", "product_name", "target_qty"}, ...]
        self.mrp_result = None
        self.products_cache = []
        self.checked_ids  = set()
        self.hist_checked = set()
        self.history_tree = None
        self.loaded_plan_id   = None   # 불러온 계획의 ID (덮어쓰기 판별용)
        self.loaded_plan_name = None   # 불러온 계획명 (기본값 표시용)

    def render(self):
        """페이지 렌더링"""
        # 스크롤 가능한 프레임
        self.scroll_frame = self.app._create_scrollable_frame()

        # ── 타이틀 ──
        title_frame = tk.Frame(self.scroll_frame, bg=COLORS["bg"])
        title_frame.pack(fill=tk.X, padx=5, pady=(5, 10))
        tk.Label(title_frame, text="📋 자재소요계획 (MRP)",
                 bg=COLORS["bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["title"], "bold")).pack(side=tk.LEFT)

        # ── 1. 생산 계획 입력 카드 ──
        self._build_plan_input_card()

        # ── 2. 생산 계획 목록 카드 ──
        self._build_plan_list_card()

        # ── 3. 소요량 계산 버튼 영역 ──
        self._build_action_bar()

        # ── 4. 결과 영역 (계산 후 표시) ──
        self.result_frame = tk.Frame(self.scroll_frame, bg=COLORS["bg"])
        self.result_frame.pack(fill=tk.X, padx=5)

        # ── 5. 저장된 MRP 계획 이력 목록 ──
        self._build_history_card()

        # 제품 목록 캐시 로드
        self._load_products_cache()
        self._load_history()

    # ═══════════════════════════════════════════
    # 1. 생산 계획 입력 카드
    # ═══════════════════════════════════════════
    def _build_plan_input_card(self):
        card = tk.Frame(self.scroll_frame, bg=COLORS["card_bg"],
                        highlightbackground=COLORS["border"], highlightthickness=1)
        card.pack(fill=tk.X, padx=5, pady=(0, 5))

        tk.Label(card, text="생산 계획 입력",
                 bg=COLORS["card_bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")).pack(
            anchor="w", padx=15, pady=(12, 5))

        tk.Label(card, text="제품을 검색하여 선택하고, 목표 수량을 입력한 뒤 추가하세요.",
                 bg=COLORS["card_bg"], fg=COLORS["text_secondary"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(anchor="w", padx=15, pady=(0, 8))

        form = tk.Frame(card, bg=COLORS["card_bg"])
        form.pack(fill=tk.X, padx=15, pady=(0, 12))

        # 제품 검색
        tk.Label(form, text="제품 검색:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"], "bold")).grid(
            row=0, column=0, sticky="e", padx=(0, 8), pady=5)

        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(form, textvariable=self.search_var,
                                     font=(FONT_FAMILY, FONT_SIZES["body"]), width=35)
        self.search_entry.grid(row=0, column=1, padx=(0, 5), pady=5, sticky="w")
        self.search_entry.bind("<KeyRelease>", self._on_search_change)
        self.search_entry.bind("<FocusIn>", self._on_search_focus)

        # 제품 정보 라벨
        self.product_info_label = tk.Label(form, text="", bg=COLORS["card_bg"],
                                           fg=COLORS["info"],
                                           font=(FONT_FAMILY, FONT_SIZES["small"]))
        self.product_info_label.grid(row=0, column=2, padx=10, sticky="w")

        # 자동완성 리스트박스 (오버레이)
        self.autocomplete_frame = tk.Frame(form, bg="white",
                                           highlightbackground=COLORS["border"],
                                           highlightthickness=1)
        self.autocomplete_listbox = tk.Listbox(self.autocomplete_frame,
                                                font=(FONT_FAMILY, FONT_SIZES["small"]),
                                                height=6, width=50,
                                                selectbackground=COLORS["primary"],
                                                selectforeground="white",
                                                activestyle="none",
                                                cursor="hand2")
        self.autocomplete_listbox.pack(fill=tk.BOTH, expand=True)
        self.autocomplete_listbox.bind("<<ListboxSelect>>", self._on_autocomplete_select)
        self.autocomplete_listbox.bind("<Double-Button-1>", self._on_autocomplete_select)

        # 목표 수량
        tk.Label(form, text="목표 수량:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"], "bold")).grid(
            row=1, column=0, sticky="e", padx=(0, 8), pady=5)

        qty_frame = tk.Frame(form, bg=COLORS["card_bg"])
        qty_frame.grid(row=1, column=1, sticky="w", pady=5)

        self.qty_entry = tk.Entry(qty_frame, font=(FONT_FAMILY, FONT_SIZES["body"]), width=12)
        self.qty_entry.pack(side=tk.LEFT, padx=(0, 10))
        self.qty_entry.bind("<Return>", lambda e: self._add_to_plan())

        # 최대 생산 가능 수량 라벨
        self.max_prod_label = tk.Label(qty_frame, text="",
                                        bg=COLORS["card_bg"], fg=COLORS["text_secondary"],
                                        font=(FONT_FAMILY, FONT_SIZES["small"]))
        self.max_prod_label.pack(side=tk.LEFT)

        # 추가 버튼
        tk.Button(form, text="➕ 계획에 추가",
                  font=(FONT_FAMILY, FONT_SIZES["body"], "bold"),
                  bg=COLORS["primary"], fg="white",
                  padx=15, pady=4, cursor="hand2",
                  command=self._add_to_plan).grid(
            row=1, column=2, padx=10, pady=5, sticky="w")

        # 선택된 제품 ID 저장용
        self.selected_product_id = None
        self.selected_product_name = None

    # ═══════════════════════════════════════════
    # 2. 생산 계획 목록 카드
    # ═══════════════════════════════════════════
    def _build_plan_list_card(self):
        card = tk.Frame(self.scroll_frame, bg=COLORS["card_bg"],
                        highlightbackground=COLORS["border"], highlightthickness=1)
        card.pack(fill=tk.X, padx=5, pady=(0, 5))

        header = tk.Frame(card, bg=COLORS["card_bg"])
        header.pack(fill=tk.X, padx=15, pady=(12, 5))

        left_hdr = tk.Frame(header, bg=COLORS["card_bg"])
        left_hdr.pack(side=tk.LEFT)
        tk.Label(left_hdr, text="생산 계획 목록",
                 bg=COLORS["card_bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")).pack(side=tk.LEFT)
        self.plan_count_label = tk.Label(left_hdr, text="(0개 제품)",
                                          bg=COLORS["card_bg"], fg=COLORS["text_secondary"],
                                          font=(FONT_FAMILY, FONT_SIZES["small"]))
        self.plan_count_label.pack(side=tk.LEFT, padx=8)
        tk.Button(left_hdr, text="전체 선택",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg="#e2e8f0", fg=COLORS["text"], padx=6, pady=1, cursor="hand2",
                  command=self._select_all_plan).pack(side=tk.LEFT, padx=(12, 2))
        tk.Button(left_hdr, text="전체 해제",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg="#e2e8f0", fg=COLORS["text"], padx=6, pady=1, cursor="hand2",
                  command=self._deselect_all_plan).pack(side=tk.LEFT, padx=2)

        right_hdr = tk.Frame(header, bg=COLORS["card_bg"])
        right_hdr.pack(side=tk.RIGHT)
        tk.Button(right_hdr, text="수량 수정",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg="#0ea5e9", fg="white", padx=8, pady=2, cursor="hand2",
                  command=self._edit_qty_dialog).pack(side=tk.LEFT, padx=2)
        tk.Button(right_hdr, text="선택 삭제",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg=COLORS["warning"], fg="white", padx=8, pady=2, cursor="hand2",
                  command=self._delete_checked_plans).pack(side=tk.LEFT, padx=2)
        tk.Button(right_hdr, text="전체 삭제",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg=COLORS["danger"], fg="white", padx=8, pady=2, cursor="hand2",
                  command=self._clear_plan).pack(side=tk.LEFT, padx=(2, 0))

        tree_frame = tk.Frame(card, bg=COLORS["card_bg"])
        tree_frame.pack(fill=tk.X, padx=15, pady=(0, 12))

        columns = ("check", "no", "product_id", "product_name", "current_stock",
                    "target_qty", "need_produce", "max_producible", "bottleneck")
        self.plan_tree = ttk.Treeview(tree_frame, columns=columns,
                                       show="headings", height=5)

        col_config = [
            ("check",          "V",            36,  "center"),
            ("no",             "No",            36,  "center"),
            ("product_id",     "제품코드",     100,  "center"),
            ("product_name",   "제품명",       150,  "w"),
            ("current_stock",  "현재재고",      80,  "center"),
            ("target_qty",     "목표수량",      80,  "center"),
            ("need_produce",   "추가생산필요", 100,  "center"),
            ("max_producible", "최대생산가능", 100,  "center"),
            ("bottleneck",     "병목부품",     150,  "w"),
        ]
        for col_id, heading, width, anchor in col_config:
            self.plan_tree.heading(col_id, text=heading)
            self.plan_tree.column(col_id, width=width, anchor=anchor,
                                   minwidth=width if col_id in ("check","no") else 30,
                                   stretch=(col_id not in ("check","no")))

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.plan_tree.yview)
        self.plan_tree.configure(yscrollcommand=scrollbar.set)
        self.plan_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def _on_plan_resize(event):
            total = self.plan_tree.winfo_width()
            fixed = 36+36+100+80+80+100+100+20
            remaining = max(200, total - fixed)
            self.plan_tree.column("product_name", width=int(remaining * 0.55))
            self.plan_tree.column("bottleneck",   width=int(remaining * 0.45))
        self.plan_tree.bind("<Configure>", _on_plan_resize)
        self.plan_tree.bind("<ButtonRelease-1>", self._on_plan_click)

        self.plan_menu = tk.Menu(self.plan_tree, tearoff=0)
        self.plan_menu.add_command(label="수량 수정",  command=self._edit_qty_dialog)
        self.plan_menu.add_separator()
        self.plan_menu.add_command(label="이 항목 삭제", command=self._delete_selected_plan)
        self.plan_tree.bind("<Button-3>", self._show_plan_menu)

    # ═══════════════════════════════════════════
    # 3. 액션 바
    # ═══════════════════════════════════════════
    def _build_action_bar(self):
        action_frame = tk.Frame(self.scroll_frame, bg=COLORS["bg"])
        action_frame.pack(fill=tk.X, padx=5, pady=5)

        left = tk.Frame(action_frame, bg=COLORS["bg"])
        left.pack(side=tk.LEFT)

        # 안전재고 반영 체크박스
        self.safety_stock_var = tk.BooleanVar(value=False)
        self.safety_check = tk.Checkbutton(
            left, text="안전재고 반영 (발주 수량에 안전재고 포함)",
            variable=self.safety_stock_var,
            bg=COLORS["bg"], fg=COLORS["text"],
            font=(FONT_FAMILY, FONT_SIZES["small"]),
            activebackground=COLORS["bg"],
            selectcolor="white")
        self.safety_check.pack(side=tk.LEFT, padx=(5, 20))

        right = tk.Frame(action_frame, bg=COLORS["bg"])
        right.pack(side=tk.RIGHT)

        # 소요량 계산 버튼
        self.calc_btn = tk.Button(right, text="🔍 소요량 계산",
                                   font=(FONT_FAMILY, FONT_SIZES["body_large"], "bold"),
                                   bg=COLORS["success"], fg="white",
                                   padx=20, pady=8, cursor="hand2",
                                   command=self._calculate_mrp)
        self.calc_btn.pack(side=tk.LEFT, padx=5)

        # 엑셀 내보내기 버튼
        self.export_btn = tk.Button(right, text="📥 엑셀 내보내기",
                                     font=(FONT_FAMILY, FONT_SIZES["body_large"], "bold"),
                                     bg=COLORS["info"], fg="white",
                                     padx=20, pady=8, cursor="hand2",
                                     state="disabled",
                                     command=self._export_excel)
        self.export_btn.pack(side=tk.LEFT, padx=5)

    # ═══════════════════════════════════════════
    # 자동완성 검색 로직
    # ═══════════════════════════════════════════
    def _load_products_cache(self):
        """제품 목록 캐시 로드 (백그라운드)"""
        def load():
            try:
                products = self.app.db.get_all_products()
                self.products_cache = products
            except:
                self.products_cache = []
        threading.Thread(target=load, daemon=True).start()

    def _on_search_focus(self, event=None):
        """검색창 포커스 시 전체 목록 표시"""
        text = self.search_var.get().strip()
        if not text:
            self._show_autocomplete(self.products_cache[:10])

    def _on_search_change(self, event=None):
        """검색어 변경 시 자동완성 업데이트"""
        # ESC 키로 닫기
        if event and event.keysym == "Escape":
            self._hide_autocomplete()
            return

        # 아래/위 화살표로 리스트 탐색
        if event and event.keysym == "Down":
            self.autocomplete_listbox.focus_set()
            if self.autocomplete_listbox.size() > 0:
                self.autocomplete_listbox.selection_set(0)
            return

        text = self.search_var.get().strip().lower()
        if not text:
            self._show_autocomplete(self.products_cache[:10])
            return

        # 제품코드 또는 제품명으로 필터링
        filtered = []
        for p in self.products_cache:
            code = str(p.get("제품코드", "")).lower()
            name = str(p.get("제품명", "")).lower()
            if text in code or text in name:
                filtered.append(p)

        self._show_autocomplete(filtered[:10])

    def _show_autocomplete(self, products):
        """자동완성 리스트 표시"""
        self.autocomplete_listbox.delete(0, tk.END)

        if not products:
            self._hide_autocomplete()
            return

        for p in products:
            code = str(p.get("제품코드", ""))
            name = str(p.get("제품명", ""))
            stock = int(p.get("현재재고", 0))
            self.autocomplete_listbox.insert(tk.END, f"{code}  |  {name}  (재고: {stock}개)")

        # 리스트 위치 설정 (검색창 바로 아래)
        self.autocomplete_frame.place(
            in_=self.search_entry,
            x=0, y=self.search_entry.winfo_height(),
            width=self.search_entry.winfo_width() + 150)

    def _hide_autocomplete(self):
        """자동완성 리스트 숨기기"""
        self.autocomplete_frame.place_forget()

    def _on_autocomplete_select(self, event=None):
        """자동완성 항목 선택"""
        selection = self.autocomplete_listbox.curselection()
        if not selection:
            return

        idx = selection[0]
        text = self.search_var.get().strip().lower()

        # 필터링된 목록에서 선택
        if text:
            filtered = [p for p in self.products_cache
                        if text in str(p.get("제품코드", "")).lower()
                        or text in str(p.get("제품명", "")).lower()]
        else:
            filtered = self.products_cache[:10]

        if idx < len(filtered):
            product = filtered[idx]
            self.selected_product_id = str(product["제품코드"])
            self.selected_product_name = str(product["제품명"])

            self.search_var.set(f"{self.selected_product_id} - {self.selected_product_name}")
            self._hide_autocomplete()

            # 제품 정보 표시
            stock = int(product.get("현재재고", 0))
            self.product_info_label.configure(
                text=f"현재재고: {stock}개")

            # 문제 3 수정: 제품 선택 시 get_max_producible() 개별 API 호출 제거
            # 최대 생산 가능 수량은 '소요량 계산' 버튼 클릭 시 일괄 계산됩니다.
            self.max_prod_label.configure(
                text="※ 최대 생산 가능 수량은 '소요량 계산' 후 확인하세요.",
                fg=COLORS["text_secondary"])

            # 수량 입력으로 포커스 이동
            self.qty_entry.focus_set()

    # ═══════════════════════════════════════════
    # 생산 계획 관리
    # ═══════════════════════════════════════════
    def _add_to_plan(self):
        """생산 계획에 제품 추가"""
        if not self.selected_product_id:
            messagebox.showwarning("알림", "제품을 먼저 검색하여 선택해 주세요.")
            self.search_entry.focus_set()
            return

        qty_text = self.qty_entry.get().strip()
        if not qty_text:
            messagebox.showwarning("알림", "목표 수량을 입력해 주세요.")
            self.qty_entry.focus_set()
            return

        try:
            target_qty = int(qty_text)
            if target_qty <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("알림", "목표 수량은 1 이상의 정수를 입력해 주세요.")
            self.qty_entry.focus_set()
            return

        # 중복 체크
        for plan in self.production_plan:
            if plan["product_id"] == self.selected_product_id:
                if messagebox.askyesno("중복 제품",
                                        f"{self.selected_product_name}이(가) 이미 목록에 있습니다.\n"
                                        f"기존 수량을 {target_qty}개로 변경하시겠습니까?"):
                    plan["target_qty"] = target_qty
                    self._refresh_plan_tree()
                return

        self.production_plan.append({
            "product_id": self.selected_product_id,
            "product_name": self.selected_product_name,
            "target_qty": target_qty,
        })

        self._refresh_plan_tree()

        # 입력 초기화
        self.search_var.set("")
        self.qty_entry.delete(0, tk.END)
        self.selected_product_id = None
        self.selected_product_name = None
        self.product_info_label.configure(text="")
        self.max_prod_label.configure(text="")
        self.search_entry.focus_set()

    def _refresh_plan_tree(self):
        """생산 계획 트리뷰 새로고침
        문제 3 수정: 계획 추가/삭제 시마다 get_max_producible()을 제품 수만큼
        반복 호출하던 구조를 제거합니다.
        → 최대생산가능/병목부품은 '소요량 계산' 버튼 클릭 후에만 표시됩니다.
          (API 호출 횟수: 제품 N개 × 2회 → 0회 로 감소)
        """
        self.plan_tree.delete(*self.plan_tree.get_children())

        products_map = {str(p["제품코드"]): p for p in self.products_cache}

        for i, plan in enumerate(self.production_plan, 1):
            pid           = plan["product_id"]
            product       = products_map.get(pid)
            current_stock = int(product["현재재고"]) if product else 0
            need          = max(0, plan["target_qty"] - current_stock)

            chk = "[v]" if pid in self.checked_ids else "[ ]"
            tag = "need" if need > 0 else "ok"
            self.plan_tree.insert("", tk.END, values=(
                chk, i, pid, plan["product_name"],
                f"{current_stock}개", f"{plan['target_qty']}개",
                f"{need}개" if need > 0 else "충분",
                "계산 전",
                "계산 전"
            ), tags=(tag,))

        self.plan_tree.tag_configure("need", foreground=COLORS["danger"])
        self.plan_tree.tag_configure("ok",   foreground=COLORS["success"])

        self.plan_count_label.configure(text=f"({len(self.production_plan)}개 제품)")

    def _delete_selected_plan(self):
        """선택된 계획 항목 삭제"""
        selected = self.plan_tree.selection()
        if not selected:
            return
        values = self.plan_tree.item(selected[0])["values"]
        product_id = str(values[2])  # check 컬럼 추가로 인덱스 2

        self.production_plan = [p for p in self.production_plan if p["product_id"] != product_id]
        self._refresh_plan_tree()

    def _show_plan_menu(self, event):
        """우클릭 메뉴"""
        item = self.plan_tree.identify_row(event.y)
        if item:
            self.plan_tree.selection_set(item)
            self.plan_menu.post(event.x_root, event.y_root)

    def _clear_plan(self):
        """전체 계획 삭제 및 초기화"""
        if not self.production_plan:
            messagebox.showinfo("알림", "계획 목록이 이미 비어 있습니다.")
            return
        if messagebox.askyesno("초기화 확인", "생산 계획 목록 전체를 삭제하고 새로 시작하시겠습니까?"):
            self.production_plan = []
            self.checked_ids.clear()
            self.mrp_result = None
            self.loaded_plan_id   = None   # ← 새로 시작 시 연결 해제
            self.loaded_plan_name = None
            self._refresh_plan_tree()
            for w in self.result_frame.winfo_children():
                w.destroy()
            self.result_frame.pack_forget()
            self.result_frame.pack(fill=tk.X, padx=5)
            self.export_btn.configure(state="disabled")
            self.app.root.update_idletasks()

    # ═══════════════════════════════════════════
    # 생산 계획 체크박스
    # ═══════════════════════════════════════════
    def _on_plan_click(self, event):
        col = self.plan_tree.identify_column(event.x)
        row = self.plan_tree.identify_row(event.y)
        if not row:
            return
        vals = self.plan_tree.item(row)["values"]
        pid  = str(vals[2])
        if col == "#1":
            if pid in self.checked_ids:
                self.checked_ids.discard(pid)
            else:
                self.checked_ids.add(pid)
            self._refresh_plan_tree()

    def _select_all_plan(self):
        self.checked_ids = {p["product_id"] for p in self.production_plan}
        self._refresh_plan_tree()

    def _deselect_all_plan(self):
        self.checked_ids.clear()
        self._refresh_plan_tree()

    def _delete_checked_plans(self):
        if not self.checked_ids:
            messagebox.showinfo("알림", "삭제할 항목을 먼저 체크해 주세요.")
            return
        names = [p["product_name"] for p in self.production_plan
                 if p["product_id"] in self.checked_ids]
        body  = "\n".join("  - " + n for n in names)
        msg   = str(len(names)) + "개 항목을 삭제하시겠습니까?\n\n" + body
        if not messagebox.askyesno("선택 삭제", msg):
            return
        self.production_plan = [
            p for p in self.production_plan
            if p["product_id"] not in self.checked_ids
        ]
        self.checked_ids.clear()
        self._refresh_plan_tree()

    def _edit_qty_dialog(self):
        selected = self.plan_tree.selection()
        if len(self.checked_ids) == 1:
            pid = next(iter(self.checked_ids))
        elif selected:
            vals = self.plan_tree.item(selected[0])["values"]
            pid  = str(vals[2])
        else:
            messagebox.showinfo("알림", "수정할 항목을 체크하거나 클릭으로 선택해 주세요.")
            return
        if len(self.checked_ids) > 1:
            messagebox.showinfo("알림", "수량 수정은 한 번에 1개 항목만 가능합니다.")
            return
        plan = next((p for p in self.production_plan if p["product_id"] == pid), None)
        if not plan:
            return

        dlg = tk.Toplevel(self.app.root)
        dlg.title("목표 수량 수정")
        self.app.center_dialog(dlg, 340, 200)
        dlg.resizable(False, False)
        dlg.transient(self.app.root)
        dlg.grab_set()

        tk.Label(dlg,
                 text=plan["product_name"] + " (" + pid + ")",
                 font=(FONT_FAMILY, FONT_SIZES["body"], "bold"),
                 padx=24, pady=16).pack(anchor="w")
        row_f = tk.Frame(dlg)
        row_f.pack(fill=tk.X, padx=24, pady=8)
        tk.Label(row_f, text="목표 수량:",
                 font=(FONT_FAMILY, FONT_SIZES["small"], "bold")).pack(side=tk.LEFT)
        ent = tk.Entry(row_f, font=(FONT_FAMILY, FONT_SIZES["body"]), width=12)
        ent.pack(side=tk.LEFT, padx=10)
        ent.insert(0, str(plan["target_qty"]))
        ent.select_range(0, tk.END)
        ent.focus_set()
        tk.Label(row_f, text="개",
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)

        def save():
            try:
                qty = int(ent.get().strip())
                if qty <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("입력 오류", "1 이상의 정수를 입력해 주세요.", parent=dlg)
                return
            plan["target_qty"] = qty
            dlg.destroy()
            self.checked_ids.clear()
            self._refresh_plan_tree()

        btn_f = tk.Frame(dlg)
        btn_f.pack(pady=12)
        qty_save_btn = tk.Button(btn_f, text="저장",
                                 font=(FONT_FAMILY, FONT_SIZES["body"], "bold"),
                                 bg=COLORS["primary"], fg="white",
                                 padx=20, pady=6, cursor="hand2",
                                 command=save)
        qty_save_btn.pack(side=tk.LEFT, padx=6)
        tk.Button(btn_f, text="취소",
                  padx=20, pady=6, cursor="hand2",
                  command=dlg.destroy).pack(side=tk.LEFT, padx=6)
        ent.bind("<Return>", lambda e: flash_btn(qty_save_btn, save))

    # ═══════════════════════════════════════════
    # MRP 계획 저장
    # ═══════════════════════════════════════════
    def _save_plan_dialog(self):
        if not self.mrp_result:
            messagebox.showwarning("알림", "먼저 소요량 계산을 실행해 주세요.")
            return

        # ── 불러온 계획이 있으면 덮어쓰기 여부 먼저 물어봄 ──
        if self.loaded_plan_id:
            ans = messagebox.askyesnocancel(
                "저장 방식 선택",
                f"'{self.loaded_plan_name}' 계획을 수정했습니다.\n\n"
                "예 → 기존 계획에 덮어쓰기\n"
                "아니오 → 새 계획으로 저장\n"
                "취소 → 돌아가기"
            )
            if ans is None:      # 취소
                return
            if ans:              # 예 → 덮어쓰기
                self._do_update(self.loaded_plan_id, self.loaded_plan_name)
                return
            # ans == False → 새 계획으로 저장 (아래 다이얼로그로 진행)

        from datetime import datetime as _dt
        default_name = (self.loaded_plan_name or "") or _dt.now().strftime("%Y-%m-%d") + " MRP 계획"

        dlg = tk.Toplevel(self.app.root)
        dlg.title("MRP 계획 저장")
        self.app.center_dialog(dlg, 380, 170)
        dlg.resizable(False, False)
        dlg.transient(self.app.root)
        dlg.grab_set()

        tk.Label(dlg, text="계획명을 입력하세요:",
                 font=(FONT_FAMILY, FONT_SIZES["body"], "bold"),
                 padx=24).pack(anchor="w", pady=(20, 4))
        ent = tk.Entry(dlg, font=(FONT_FAMILY, FONT_SIZES["body"]), width=34)
        ent.pack(padx=24, pady=(0, 4))
        ent.insert(0, default_name)
        ent.select_range(0, tk.END)
        ent.focus_set()

        def do_save():
            plan_name = ent.get().strip()
            if not plan_name:
                messagebox.showwarning("입력 오류", "계획명을 입력해 주세요.", parent=dlg)
                return
            dlg.destroy()
            self._do_save(plan_name)

        btn_f = tk.Frame(dlg)
        btn_f.pack(pady=10)
        plan_save_btn = tk.Button(btn_f, text="저장",
                                  font=(FONT_FAMILY, FONT_SIZES["body"], "bold"),
                                  bg=COLORS["primary"], fg="white", padx=18, pady=5,
                                  cursor="hand2", command=do_save)
        plan_save_btn.pack(side=tk.LEFT, padx=6)
        tk.Button(btn_f, text="취소",
                  padx=18, pady=5, cursor="hand2",
                  command=dlg.destroy).pack(side=tk.LEFT, padx=6)
        ent.bind("<Return>", lambda e: flash_btn(plan_save_btn, do_save))

    def _do_save(self, plan_name):
        def worker():
            try:
                plan_id = self.app.db.save_mrp_plan(
                    plan_name, self.production_plan,
                    self.mrp_result, self.safety_stock_var.get())
                # 새로 저장된 계획을 현재 loaded 계획으로 설정 (이후 다시 저장 시 덮어쓰기 가능)
                self.loaded_plan_id   = plan_id
                self.loaded_plan_name = plan_name
                ok_msg = "'" + plan_name + "' 저장 완료 (ID: " + plan_id + ")"
                self.app.root.after(0, lambda: (
                    messagebox.showinfo("저장 완료", ok_msg),
                    self._load_history()
                ))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("저장 오류", err))
        threading.Thread(target=worker, daemon=True).start()

    def _do_update(self, plan_id, plan_name):
        """기존 MRP 계획 덮어쓰기"""
        def worker():
            try:
                ok = self.app.db.update_mrp_plan(
                    plan_id, plan_name, self.production_plan,
                    self.mrp_result, self.safety_stock_var.get())
                if ok:
                    ok_msg = "'" + plan_name + "' 계획이 업데이트되었습니다."
                    self.app.root.after(0, lambda: (
                        messagebox.showinfo("저장 완료", ok_msg),
                        self._load_history()
                    ))
                else:
                    # plan_id 가 시트에 없는 경우 → 새 계획으로 저장 제안
                    self.app.root.after(0, lambda: self._offer_save_as_new(plan_name))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror(
                    "저장 오류", f"계획 업데이트 중 오류가 발생했습니다:\n{err}"))
        threading.Thread(target=worker, daemon=True).start()

    def _offer_save_as_new(self, plan_name):
        """덮어쓰기 실패 시 새 계획으로 저장할지 물어봄"""
        ans = messagebox.askyesno(
            "저장 오류",
            f"'{plan_name}' 계획을 시트에서 찾을 수 없습니다.\n"
            "(삭제됐거나 ID가 변경됐을 수 있습니다.)\n\n"
            "새 계획으로 저장하시겠습니까?"
        )
        if ans:
            self._do_save(plan_name)

    # ═══════════════════════════════════════════
    # 저장된 MRP 계획 이력 목록 카드
    # ═══════════════════════════════════════════
    def _build_history_card(self):
        sep = tk.Frame(self.scroll_frame, bg=COLORS["border"], height=2)
        sep.pack(fill=tk.X, padx=5, pady=(16, 0))

        card = tk.Frame(self.scroll_frame, bg=COLORS["card_bg"],
                        highlightbackground=COLORS["border"], highlightthickness=1)
        card.pack(fill=tk.X, padx=5, pady=(0, 12))

        hdr = tk.Frame(card, bg=COLORS["card_bg"])
        hdr.pack(fill=tk.X, padx=15, pady=(12, 6))

        left_h = tk.Frame(hdr, bg=COLORS["card_bg"])
        left_h.pack(side=tk.LEFT)
        tk.Label(left_h, text="저장된 MRP 계획 목록",
                 bg=COLORS["card_bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")).pack(side=tk.LEFT)
        self._hist_count_lbl = tk.Label(left_h, text="(0건)",
                                         bg=COLORS["card_bg"], fg=COLORS["text_secondary"],
                                         font=(FONT_FAMILY, FONT_SIZES["small"]))
        self._hist_count_lbl.pack(side=tk.LEFT, padx=8)

        right_h = tk.Frame(hdr, bg=COLORS["card_bg"])
        right_h.pack(side=tk.RIGHT)
        tk.Button(right_h, text="불러오기",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg=COLORS["primary"], fg="white", padx=8, pady=2, cursor="hand2",
                  command=self._load_checked_history).pack(side=tk.LEFT, padx=2)
        tk.Button(right_h, text="이름 수정",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg="#0ea5e9", fg="white", padx=8, pady=2, cursor="hand2",
                  command=self._rename_history).pack(side=tk.LEFT, padx=2)
        tk.Button(right_h, text="선택 삭제",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg=COLORS["danger"], fg="white", padx=8, pady=2, cursor="hand2",
                  command=self._delete_checked_history).pack(side=tk.LEFT, padx=2)
        tk.Button(right_h, text="새로고침",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg="#e2e8f0", fg=COLORS["text"], padx=8, pady=2, cursor="hand2",
                  command=self._load_history).pack(side=tk.LEFT, padx=(2, 0))

        tf = tk.Frame(card, bg=COLORS["card_bg"])
        tf.pack(fill=tk.X, padx=15, pady=(0, 12))

        hist_cols = ("hchk", "hno", "hid", "hdate", "hname", "hcnt", "hsafe")
        self.history_tree = ttk.Treeview(tf, columns=hist_cols,
                                          show="headings", height=6)
        for cid, heading, width, anchor in [
            ("hchk",  "V",        36,  "center"),
            ("hno",   "No",        36,  "center"),
            ("hid",   "저장ID",    70,  "center"),
            ("hdate", "저장일시", 130,  "center"),
            ("hname", "계획명",   220,  "w"),
            ("hcnt",  "제품수",    60,  "center"),
            ("hsafe", "안전재고",  70,  "center"),
        ]:
            self.history_tree.heading(cid, text=heading)
            self.history_tree.column(cid, width=width, anchor=anchor,
                                      minwidth=width if cid in ("hchk", "hno") else 30,
                                      stretch=(cid not in ("hchk", "hno")))

        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=vsb.set)
        self.history_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        self.history_tree.bind("<ButtonRelease-1>", self._on_hist_click)
        self.history_tree.bind("<Double-1>",        lambda e: self._load_checked_history())

        self._hist_menu = tk.Menu(self.history_tree, tearoff=0)
        self._hist_menu.add_command(label="불러오기",  command=self._load_checked_history)
        self._hist_menu.add_command(label="이름 수정", command=self._rename_history)
        self._hist_menu.add_separator()
        self._hist_menu.add_command(label="삭제",      command=self._delete_checked_history)
        self.history_tree.bind("<Button-3>", self._hist_right_click)

    def _load_history(self):
        def load():
            try:
                rows = self.app.db.get_all_mrp_history()
                self.app.root.after(0, lambda: self._render_history(rows))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err))
        threading.Thread(target=load, daemon=True).start()

    def _render_history(self, rows):
        if not self.history_tree:
            return
        self.history_tree.delete(*self.history_tree.get_children())
        for i, row in enumerate(rows, 1):
            pid  = str(row.get("저장ID", ""))
            chk  = "[v]" if pid in self.hist_checked else "[ ]"
            safe = "Y" if str(row.get("안전재고반영", "FALSE")).upper() == "TRUE" else "N"
            self.history_tree.insert("", tk.END, values=(
                chk, i, pid,
                row.get("저장일시", ""),
                row.get("계획명", ""),
                str(row.get("제품수", "")) + "개",
                safe,
            ))
        self._hist_count_lbl.configure(text="(" + str(len(rows)) + "건)")

    def _on_hist_click(self, event):
        col = self.history_tree.identify_column(event.x)
        row = self.history_tree.identify_row(event.y)
        if not row:
            return
        vals = self.history_tree.item(row)["values"]
        pid  = str(vals[2])
        if col == "#1":
            if pid in self.hist_checked:
                self.hist_checked.discard(pid)
            else:
                self.hist_checked.add(pid)
        else:
            self.hist_checked = {pid}
        self._refresh_history_checks()

    def _refresh_history_checks(self):
        for item in self.history_tree.get_children():
            vals      = list(self.history_tree.item(item)["values"])
            vals[0]   = "[v]" if str(vals[2]) in self.hist_checked else "[ ]"
            self.history_tree.item(item, values=vals)

    def _hist_right_click(self, event):
        row = self.history_tree.identify_row(event.y)
        if row:
            vals = self.history_tree.item(row)["values"]
            self.hist_checked = {str(vals[2])}
            self._refresh_history_checks()
            self._hist_menu.post(event.x_root, event.y_root)

    def _load_checked_history(self):
        if not self.hist_checked:
            messagebox.showinfo("알림", "불러올 계획을 선택해 주세요.")
            return
        if len(self.hist_checked) > 1:
            messagebox.showinfo("알림", "불러오기는 한 번에 1개만 가능합니다.")
            return
        plan_id = next(iter(self.hist_checked))
        def load():
            try:
                detail = self.app.db.get_mrp_plan_detail(plan_id)
                self.app.root.after(0, lambda: self._apply_loaded_plan(detail))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err))
        threading.Thread(target=load, daemon=True).start()

    def _apply_loaded_plan(self, detail):
        if not detail:
            messagebox.showerror("오류", "계획 데이터를 불러올 수 없습니다.")
            return
        self.production_plan = detail.get("production_plan", [])
        self.checked_ids.clear()
        self.mrp_result = None
        self.loaded_plan_id   = detail.get("plan_id")   # ← 불러온 계획 ID 기억
        self.loaded_plan_name = detail.get("name", "")  # ← 불러온 계획명 기억
        self.safety_stock_var.set(detail.get("include_safety", False))
        # 결과 영역 초기화
        for w in self.result_frame.winfo_children():
            w.destroy()
        self.result_frame.pack_forget()
        self.result_frame.pack(fill=tk.X, padx=5)
        self.export_btn.configure(state="disabled")
        self.app.root.update_idletasks()
        self._refresh_plan_tree()
        name = detail.get("name", "")
        messagebox.showinfo("불러오기 완료",
                            "'" + name + "' 계획이 복원되었습니다.\n"
                            "수정 후 [소요량 계산] 버튼을 눌러 계산하세요.")

    def _rename_history(self):
        if not self.hist_checked:
            messagebox.showinfo("알림", "수정할 계획을 선택해 주세요.")
            return
        if len(self.hist_checked) > 1:
            messagebox.showinfo("알림", "이름 수정은 한 번에 1개만 가능합니다.")
            return
        plan_id  = next(iter(self.hist_checked))
        cur_name = ""
        for item in self.history_tree.get_children():
            vals = self.history_tree.item(item)["values"]
            if str(vals[2]) == plan_id:
                cur_name = str(vals[4])
                break

        dlg = tk.Toplevel(self.app.root)
        dlg.title("계획명 수정")
        self.app.center_dialog(dlg, 360, 160)
        dlg.resizable(False, False)
        dlg.transient(self.app.root)
        dlg.grab_set()

        tk.Label(dlg, text="새 계획명:",
                 font=(FONT_FAMILY, FONT_SIZES["body"], "bold"),
                 padx=24).pack(anchor="w", pady=(18, 4))
        ent = tk.Entry(dlg, font=(FONT_FAMILY, FONT_SIZES["body"]), width=34)
        ent.pack(padx=24)
        ent.insert(0, cur_name)
        ent.select_range(0, tk.END)
        ent.focus_set()

        def do_rename():
            new_name = ent.get().strip()
            if not new_name:
                messagebox.showwarning("입력 오류", "계획명을 입력해 주세요.", parent=dlg)
                return
            dlg.destroy()
            def worker():
                try:
                    self.app.db.update_mrp_plan_name(plan_id, new_name)
                    self.app.root.after(0, self._load_history)
                except Exception as e:
                    err = str(e)
                    self.app.root.after(0, lambda: messagebox.showerror("오류", err))
            threading.Thread(target=worker, daemon=True).start()

        btn_f = tk.Frame(dlg)
        btn_f.pack(pady=10)
        rename_btn = tk.Button(btn_f, text="저장",
                               font=(FONT_FAMILY, FONT_SIZES["body"], "bold"),
                               bg=COLORS["primary"], fg="white", padx=18, pady=5,
                               cursor="hand2", command=do_rename)
        rename_btn.pack(side=tk.LEFT, padx=6)
        tk.Button(btn_f, text="취소",
                  padx=18, pady=5, cursor="hand2",
                  command=dlg.destroy).pack(side=tk.LEFT, padx=6)
        ent.bind("<Return>", lambda e: flash_btn(rename_btn, do_rename))

    def _delete_checked_history(self):
        if not self.hist_checked:
            messagebox.showinfo("알림", "삭제할 계획을 선택해 주세요.")
            return
        n   = len(self.hist_checked)
        msg = str(n) + "건의 MRP 계획을 삭제하시겠습니까?"
        if not messagebox.askyesno("삭제 확인", msg):
            return
        ids = list(self.hist_checked)
        def worker():
            try:
                cnt = self.app.db.delete_mrp_plans(ids)
                self.hist_checked.clear()
                done = str(cnt) + "건 삭제되었습니다."
                self.app.root.after(0, lambda: (
                    messagebox.showinfo("삭제 완료", done),
                    self._load_history()
                ))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err))
        threading.Thread(target=worker, daemon=True).start()

    # ═══════════════════════════════════════════
    # MRP 계산
    # ═══════════════════════════════════════════
    def _calculate_mrp(self):
        """소요량 계산 실행"""
        if not self.production_plan:
            messagebox.showwarning("알림", "생산 계획을 먼저 입력해 주세요.")
            return

        self.calc_btn.configure(state="disabled", text="계산 중...")

        include_safety = self.safety_stock_var.get()

        def process():
            try:
                result = self.app.db.calculate_mrp(
                    self.production_plan,
                    include_safety_stock=include_safety
                )
                self.app.root.after(0, lambda: self._show_mrp_result(result))
            except Exception as e:
                self.app.root.after(0, lambda: self._show_mrp_error(str(e)))

        threading.Thread(target=process, daemon=True).start()

    def _show_mrp_error(self, error_msg):
        """계산 오류 표시"""
        self.calc_btn.configure(state="normal", text="🔍 소요량 계산")
        messagebox.showerror("계산 오류", f"MRP 계산 중 오류가 발생했습니다:\n{error_msg}")

    def _show_mrp_result(self, result):
        """MRP 계산 결과 표시"""
        self.calc_btn.configure(state="normal", text="🔍 소요량 계산")
        self.mrp_result = result
        self.export_btn.configure(state="normal")

        # 결과 영역 초기화 후 expand=True로 재팩
        for w in self.result_frame.winfo_children():
            w.destroy()
        self.result_frame.pack_forget()
        self.result_frame.pack(fill=tk.BOTH, expand=True, padx=5)

        # ── 생산 계획 요약 업데이트 ──
        self._refresh_plan_tree_with_result(result["plan_summary"])

        # ── 저장 버튼 ──
        save_bar = tk.Frame(self.result_frame, bg=COLORS["bg"])
        save_bar.pack(fill=tk.X, pady=(4, 0))
        tk.Button(save_bar, text="계획 저장",
                  font=(FONT_FAMILY, FONT_SIZES["body"], "bold"),
                  bg="#059669", fg="white", padx=16, pady=5,
                  cursor="hand2",
                  command=self._save_plan_dialog).pack(side=tk.RIGHT, padx=2)

        # ── 요약 통계 카드 ──
        stats_frame = tk.Frame(self.result_frame, bg=COLORS["bg"])
        stats_frame.pack(fill=tk.X, pady=(5, 5))

        total_parts = len(result["parts_requirement"])
        order_items = result["total_order_items"]
        order_qty = result["total_order_qty"]
        sufficient = total_parts - order_items

        stats = [
            ("총 소요 부품", f"{total_parts}종", COLORS["info"]),
            ("재고 충분", f"{sufficient}종", COLORS["success"]),
            ("발주 필요", f"{order_items}종", COLORS["danger"] if order_items > 0 else COLORS["success"]),
            ("총 발주 수량", f"{order_qty:,}개", COLORS["warning"] if order_qty > 0 else COLORS["success"]),
        ]

        for i, (label, value, color) in enumerate(stats):
            stat_card = tk.Frame(stats_frame, bg="white",
                                  highlightbackground=color, highlightthickness=2)
            stat_card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=3)
            tk.Label(stat_card, text=label, bg="white", fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["small"])).pack(pady=(8, 2))
            tk.Label(stat_card, text=value, bg="white", fg=color,
                     font=(FONT_FAMILY, FONT_SIZES["stat"], "bold")).pack(pady=(0, 8))

        # ── 발주 리스트 카드 ──
        self._build_requirement_table(result)

    def _refresh_plan_tree_with_result(self, plan_summary):
        """MRP 계산 결과로 생산 계획 트리 업데이트
        문제 3 수정: 최대생산가능/병목부품은 여기서 한 번만 채워집니다.
        calculate_mrp() 안에서 이미 일괄 계산된 결과를 받아 표시하므로
        추가 API 호출이 전혀 발생하지 않습니다.
        """
        self.plan_tree.delete(*self.plan_tree.get_children())

        for i, item in enumerate(plan_summary, 1):
            need = item["need_to_produce"]
            pid  = item["product_id"]
            chk  = "[v]" if pid in self.checked_ids else "[ ]"
            tag  = "need" if need > 0 else "ok"
            self.plan_tree.insert("", tk.END, values=(
                chk, i, pid, item["product_name"],
                f"{item['current_stock']}개", f"{item['target_qty']}개",
                f"{need}개" if need > 0 else "충분",
                f"{item['max_producible']}개",
                item["bottleneck"] if need > 0 else "-"
            ), tags=(tag,))

        self.plan_tree.tag_configure("need", foreground=COLORS["danger"])
        self.plan_tree.tag_configure("ok",   foreground=COLORS["success"])

    def _build_requirement_table(self, result):
        """발주 리스트 테이블 구성"""
        card = tk.Frame(self.result_frame, bg=COLORS["card_bg"],
                        highlightbackground=COLORS["border"], highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=0, pady=(5, 10))

        header = tk.Frame(card, bg=COLORS["card_bg"])
        header.pack(fill=tk.X, padx=15, pady=(12, 5))

        tk.Label(header, text="부품 소요량 및 발주 리스트",
                 bg=COLORS["card_bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")).pack(side=tk.LEFT)

        safety_text = " (안전재고 반영)" if self.safety_stock_var.get() else ""
        tk.Label(header, text=safety_text,
                 bg=COLORS["card_bg"], fg=COLORS["info"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT, padx=5)

        # 필터 옵션
        filter_frame = tk.Frame(card, bg=COLORS["card_bg"])
        filter_frame.pack(fill=tk.X, padx=15, pady=(0, 5))

        self.filter_var = tk.StringVar(value="all")
        tk.Radiobutton(filter_frame, text="전체 부품", variable=self.filter_var,
                       value="all", bg=COLORS["card_bg"],
                       font=(FONT_FAMILY, FONT_SIZES["small"]),
                       command=lambda: self._apply_filter(result)).pack(side=tk.LEFT, padx=(0, 15))
        tk.Radiobutton(filter_frame, text="발주 필요 부품만", variable=self.filter_var,
                       value="order_only", bg=COLORS["card_bg"],
                       font=(FONT_FAMILY, FONT_SIZES["small"]),
                       command=lambda: self._apply_filter(result)).pack(side=tk.LEFT)

        # 트리뷰
        tree_frame = tk.Frame(card, bg=COLORS["card_bg"])
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 12))

        columns = ("no", "part_id", "part_name", "supplier", "unit",
                    "total_required", "current_stock", "safety_stock",
                    "shortage", "status")
        self.req_tree = ttk.Treeview(tree_frame, columns=columns,
                                      show="headings", height=15)

        col_config = [
            ("no", "No", 40, "center"),
            ("part_id", "품번", 100, "center"),
            ("part_name", "부품명", 180, "w"),
            ("supplier", "업체명", 100, "w"),
            ("unit", "단위", 50, "center"),
            ("total_required", "총소요량", 80, "center"),
            ("current_stock", "현재재고", 80, "center"),
            ("safety_stock", "안전재고", 80, "center"),
            ("shortage", "부족수량", 80, "center"),
            ("status", "발주필요", 80, "center"),
        ]

        for col_id, heading, width, anchor in col_config:
            self.req_tree.heading(col_id, text=heading)
            self.req_tree.column(col_id, width=width, anchor=anchor)

        scrollbar_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.req_tree.yview)
        self.req_tree.configure(yscrollcommand=scrollbar_y.set)
        self.req_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        # 반응형: 부품명 컬럼 자동 조정
        def _on_req_resize(event):
            total = self.req_tree.winfo_width()
            fixed = 40 + 100 + 100 + 50 + 80 + 80 + 80 + 80 + 80 + 20  # 부품명 제외 고정 열
            remaining = max(120, total - fixed)
            self.req_tree.column("part_name", width=remaining)
        self.req_tree.bind("<Configure>", _on_req_resize)

        self.req_tree.tag_configure("order", background="#fef2f2", foreground=COLORS["danger"])
        self.req_tree.tag_configure("ok", foreground=COLORS["success"])

        self._apply_filter(result)

    def _apply_filter(self, result):
        """필터 적용"""
        self.req_tree.delete(*self.req_tree.get_children())

        filter_mode = self.filter_var.get()
        parts = result["parts_requirement"]

        for i, part in enumerate(parts, 1):
            if filter_mode == "order_only" and not part["order_needed"]:
                continue

            tag = "order" if part["order_needed"] else "ok"
            status = f"발주 {int(part['shortage'])}개" if part["order_needed"] else "충분"

            self.req_tree.insert("", tk.END, values=(
                i, part["part_id"], part["part_name"],
                part["supplier"], part["unit"],
                part["total_required"], part["current_stock"],
                part["safety_stock"], int(part["shortage"]),
                status
            ), tags=(tag,))

    # ═══════════════════════════════════════════
    # 엑셀 내보내기
    # ═══════════════════════════════════════════
    def _export_excel(self):
        """MRP 결과를 엑셀로 내보내기"""
        if not self.mrp_result:
            messagebox.showwarning("알림", "먼저 소요량 계산을 실행해 주세요.")
            return

        # 저장 경로 선택
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"MRP_발주리스트_{now}.xlsx"

        file_path = filedialog.asksaveasfilename(
            title="MRP 발주 리스트 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=default_name,
        )

        if not file_path:
            return

        try:
            self._generate_excel(file_path)
            messagebox.showinfo("저장 완료",
                                f"MRP 발주 리스트가 저장되었습니다.\n\n{file_path}")
            # 파일 열기
            os.startfile(file_path)
        except Exception as e:
            messagebox.showerror("저장 오류", f"엑셀 저장 중 오류가 발생했습니다:\n{e}")

    def _generate_excel(self, file_path):
        """엑셀 파일 생성"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        # ── Sheet 1: 생산 계획 요약 ──
        ws1 = wb.active
        ws1.title = "생산계획"

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
        normal_font = Font(name="맑은 고딕", size=10)
        title_font = Font(name="맑은 고딕", bold=True, size=14)
        danger_font = Font(name="맑은 고딕", bold=True, size=10, color="DC2626")
        success_font = Font(name="맑은 고딕", size=10, color="16A34A")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))

        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws1.merge_cells("A1:H1")
        ws1["A1"] = f"자재소요계획 (MRP) - {now}"
        ws1["A1"].font = title_font
        ws1.row_dimensions[1].height = 30

        safety_text = "안전재고 반영: 예" if self.safety_stock_var.get() else "안전재고 반영: 아니오"
        ws1.merge_cells("A2:H2")
        ws1["A2"] = safety_text
        ws1["A2"].font = Font(name="맑은 고딕", size=9, color="666666")

        headers1 = ["No", "제품코드", "제품명", "현재재고", "목표수량",
                     "추가생산필요", "최대생산가능", "병목부품"]
        for col, h in enumerate(headers1, 1):
            cell = ws1.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for i, item in enumerate(self.mrp_result["plan_summary"], 1):
            row = i + 4
            values = [i, item["product_id"], item["product_name"],
                      item["current_stock"], item["target_qty"],
                      item["need_to_produce"], item["max_producible"],
                      item["bottleneck"]]
            for col, val in enumerate(values, 1):
                cell = ws1.cell(row=row, column=col, value=val)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if col == 6 and item["need_to_produce"] > 0:
                    cell.font = danger_font

        # 열 너비
        widths1 = [6, 15, 25, 12, 12, 14, 14, 25]
        for i, w in enumerate(widths1, 1):
            ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # ── Sheet 2: 전체 발주 리스트 ──
        ws2 = wb.create_sheet("발주리스트_전체")

        ws2.merge_cells("A1:J1")
        ws2["A1"] = f"부품 소요량 및 발주 리스트 - {now}"
        ws2["A1"].font = title_font
        ws2.row_dimensions[1].height = 30

        summary = (f"총 소요 부품: {len(self.mrp_result['parts_requirement'])}종  |  "
                   f"발주 필요: {self.mrp_result['total_order_items']}종  |  "
                   f"총 발주 수량: {self.mrp_result['total_order_qty']:,}개")
        ws2.merge_cells("A2:J2")
        ws2["A2"] = summary
        ws2["A2"].font = Font(name="맑은 고딕", size=9, color="666666")

        headers2 = ["No", "품번", "부품명", "업체명", "단위",
                     "총소요량", "현재재고", "안전재고", "부족수량", "발주필요"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        order_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

        for i, part in enumerate(self.mrp_result["parts_requirement"], 1):
            row = i + 4
            status = f"발주 {int(part['shortage'])}개" if part["order_needed"] else "충분"
            values = [i, part["part_id"], part["part_name"], part["supplier"],
                      part["unit"], part["total_required"], part["current_stock"],
                      part["safety_stock"], int(part["shortage"]), status]
            for col, val in enumerate(values, 1):
                cell = ws2.cell(row=row, column=col, value=val)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if part["order_needed"]:
                    cell.fill = order_fill
                    if col in (9, 10):
                        cell.font = danger_font

        widths2 = [6, 15, 30, 15, 8, 12, 12, 12, 12, 14]
        for i, w in enumerate(widths2, 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # ── Sheet 3: 전체 부품 (업체별 그룹핑, 재고 충분 포함) ──
        ws3 = wb.create_sheet("발주현황_업체별")

        ws3.merge_cells("A1:I1")
        ws3["A1"] = f"부품 발주 현황 (업체별 전체) - {now}"
        ws3["A1"].font = title_font
        ws3.row_dimensions[1].height = 30

        # 전체 부품을 업체별로 그룹핑 (발주 필요 + 재고 충분 모두 포함)
        all_req_parts = self.mrp_result["parts_requirement"]
        suppliers = {}
        for part in all_req_parts:
            supplier = part["supplier"] if part["supplier"] else "(업체 미지정)"
            if supplier not in suppliers:
                suppliers[supplier] = []
            suppliers[supplier].append(part)

        def _qty_or_blank(val):
            """0이면 빈칸, 아니면 그대로 반환"""
            return "" if val == 0 else val

        ok_fill   = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")  # 연초록 - 충분
        need_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")  # 연빨강 - 발주필요
        ok_font   = Font(name="맑은 고딕", size=10, color="16A34A")   # 초록 글씨
        ok_bold   = Font(name="맑은 고딕", size=10, bold=True, color="16A34A")

        current_row = 3
        for supplier, parts in sorted(suppliers.items()):
            order_cnt = sum(1 for p in parts if p["order_needed"])
            ok_cnt    = len(parts) - order_cnt

            # 업체명 헤더 (발주 필요/충분 건수 표시)
            ws3.merge_cells(f"A{current_row}:I{current_row}")
            cell = ws3.cell(row=current_row, column=1,
                            value=f"▶ {supplier}  (전체 {len(parts)}종 | 발주필요 {order_cnt}종 | 충분 {ok_cnt}종)")
            cell.font = Font(name="맑은 고딕", bold=True, size=11, color="1E293B")
            cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            current_row += 1

            # 컬럼 헤더
            sub_headers = ["No", "품번", "부품명", "단위", "총소요량",
                           "현재재고", "부족수량/상태", "사용제품코드", "비고"]
            for col, h in enumerate(sub_headers, 1):
                cell = ws3.cell(row=current_row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            current_row += 1

            # 발주필요 먼저, 충분 나중 순으로 정렬
            sorted_parts = sorted(parts, key=lambda p: (not p["order_needed"], p["part_id"]))

            for j, part in enumerate(sorted_parts, 1):
                product_codes_str = ", ".join(part.get("product_codes", []))
                is_ok = not part["order_needed"]

                if is_ok:
                    shortage_val = "충분"
                else:
                    shortage_val = _qty_or_blank(int(part["shortage"]))

                values = [j,
                          part["part_id"],
                          part["part_name"],
                          part["unit"],
                          _qty_or_blank(part["total_required"]),
                          _qty_or_blank(part["current_stock"]),
                          shortage_val,
                          product_codes_str,
                          ""]

                row_fill = ok_fill if is_ok else need_fill

                for col, val in enumerate(values, 1):
                    cell = ws3.cell(row=current_row, column=col, value=val)
                    cell.fill = row_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

                    if col == 7:
                        if is_ok:
                            cell.font = ok_bold
                        elif val != "":
                            cell.font = danger_font
                        else:
                            cell.font = normal_font
                    elif col == 8:
                        cell.font = normal_font
                        cell.alignment = Alignment(horizontal="left")
                    else:
                        cell.font = ok_font if is_ok else normal_font

                current_row += 1

            current_row += 1  # 업체 간 빈 줄

        widths3 = [6, 15, 30, 8, 12, 12, 14, 30, 20]
        for i, w in enumerate(widths3, 1):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # ── Sheet 4: Raw 데이터 (피벗테이블용) ──
        ws4 = wb.create_sheet("Raw데이터")

        raw_header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
        raw_headers = ["사용제품코드", "품번", "업체명", "부품명", "재질/규격", "단가", "소요량"]
        for col, h in enumerate(raw_headers, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = raw_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws4.row_dimensions[1].height = 26

        price_font = Font(name="맑은 고딕", size=10)
        for row_idx, part in enumerate(self.mrp_result["parts_requirement"], 2):
            product_codes_str = ", ".join(part.get("product_codes", []))
            unit_price = part.get("unit_price", 0)
            values = [
                product_codes_str,
                part["part_id"],
                part["supplier"],
                part["part_name"],
                part.get("spec", ""),
                unit_price if unit_price else "",
                part["total_required"],
            ]
            for col, val in enumerate(values, 1):
                cell = ws4.cell(row=row_idx, column=col, value=val)
                cell.font = price_font
                cell.border = thin_border
                # 단가: 오른쪽 정렬 + 숫자 서식
                if col == 6 and val != "":
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0"
                # 소요량: 가운데 정렬
                elif col == 7:
                    cell.alignment = Alignment(horizontal="center")
                # 나머지: 왼쪽 정렬
                else:
                    cell.alignment = Alignment(horizontal="left")

        widths4 = [30, 15, 15, 30, 20, 14, 10]
        for i, w in enumerate(widths4, 1):
            ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # 자동 필터 적용 (피벗 활용 편의)
        ws4.auto_filter.ref = f"A1:G{len(self.mrp_result['parts_requirement']) + 1}"

        wb.save(file_path)
