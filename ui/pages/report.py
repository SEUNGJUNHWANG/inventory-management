"""
재고관리 시스템 - 엑셀 리포트 다운로드 페이지
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import threading
from datetime import datetime
from core.constants import COLORS, FONT_FAMILY, FONT_SIZES


class ReportPage:
    def __init__(self, app):
        self.app = app
        self.report_status = None

    def render(self):
        card = self.app._create_card("📊 엑셀 리포트 다운로드")

        tk.Label(card, text="현재 재고 현황 및 입출고 이력을 엑셀 파일로 다운로드합니다.",
                 bg=COLORS["card_bg"], fg=COLORS["text_secondary"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(anchor="w", padx=20, pady=(0, 15))

        options_frame = tk.Frame(card, bg=COLORS["card_bg"])
        options_frame.pack(fill=tk.X, padx=20, pady=5)

        tk.Button(options_frame, text="📋 부품 재고 현황 다운로드",
                  font=(FONT_FAMILY, FONT_SIZES["body_large"]), bg=COLORS["primary"], fg="white",
                  padx=20, pady=10, cursor="hand2",
                  command=lambda: self._download_report("parts")).pack(fill=tk.X, pady=5)

        tk.Button(options_frame, text="📦 제품 재고 현황 다운로드",
                  font=(FONT_FAMILY, FONT_SIZES["body_large"]), bg=COLORS["success"], fg="white",
                  padx=20, pady=10, cursor="hand2",
                  command=lambda: self._download_report("products")).pack(fill=tk.X, pady=5)

        tk.Button(options_frame, text="📜 입출고 이력 다운로드",
                  font=(FONT_FAMILY, FONT_SIZES["body_large"]), bg=COLORS["warning"], fg="white",
                  padx=20, pady=10, cursor="hand2",
                  command=lambda: self._download_report("history")).pack(fill=tk.X, pady=5)

        tk.Button(options_frame, text="📊 전체 리포트 다운로드 (모든 시트 포함)",
                  font=(FONT_FAMILY, FONT_SIZES["body_large"], "bold"), bg="#6366f1", fg="white",
                  padx=20, pady=10, cursor="hand2",
                  command=lambda: self._download_report("all")).pack(fill=tk.X, pady=5)

        self.report_status = tk.Label(card, text="", bg=COLORS["card_bg"],
                                      font=(FONT_FAMILY, FONT_SIZES["small"]))
        self.report_status.pack(anchor="w", padx=20, pady=15)

    def _download_report(self, report_type):
        filepath = filedialog.asksaveasfilename(
            title="리포트 저장 위치 선택",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=f"재고관리_리포트_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if not filepath:
            return

        self.report_status.configure(text="리포트 생성 중...", fg=COLORS["warning"])
        self.app.root.update()

        def generate():
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                wb = openpyxl.Workbook()
                header_font = Font(name=FONT_FAMILY, bold=True, size=11, color="FFFFFF")
                header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
                header_align = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"),
                )
                warning_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

                if report_type in ("parts", "all"):
                    ws = wb.active if report_type == "parts" else wb.create_sheet()
                    ws.title = "부품재고현황"
                    headers = ["품번", "부품명", "규격", "단위", "현재재고", "안전재고", "상태", "비고"]
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                        cell.border = thin_border

                    parts = self.app.db.get_all_parts()
                    for row, p in enumerate(parts, 2):
                        current = int(p.get("현재재고", 0))
                        safety = int(p.get("안전재고", 0))
                        status = "정상"
                        if safety > 0 and current <= safety:
                            status = "부족"
                        elif safety > 0 and current <= safety * 1.2:
                            status = "주의"

                        data = [p.get("품번", ""), p.get("부품명", ""), p.get("규격", ""),
                                p.get("단위", ""), current, safety, status, p.get("비고", "")]
                        for col, val in enumerate(data, 1):
                            cell = ws.cell(row=row, column=col, value=val)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center")
                            if status == "부족":
                                cell.fill = warning_fill

                    for col in range(1, len(headers) + 1):
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

                if report_type in ("products", "all"):
                    ws = wb.active if report_type == "products" else wb.create_sheet()
                    ws.title = "제품재고현황"
                    headers = ["제품코드", "제품명", "규격", "현재재고", "비고"]
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                        cell.border = thin_border

                    products = self.app.db.get_all_products()
                    for row, p in enumerate(products, 2):
                        data = [p.get("제품코드", ""), p.get("제품명", ""), p.get("규격", ""),
                                int(p.get("현재재고", 0)), p.get("비고", "")]
                        for col, val in enumerate(data, 1):
                            cell = ws.cell(row=row, column=col, value=val)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center")

                    for col in range(1, len(headers) + 1):
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

                if report_type in ("history", "all"):
                    ws = wb.active if report_type == "history" else wb.create_sheet()
                    ws.title = "입출고이력"
                    headers = ["일시", "구분", "유형", "품번/제품코드", "품명", "수량", "잔여재고", "관련제품", "비고"]
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                        cell.border = thin_border

                    history = self.app.db.get_all_history()
                    for row, h in enumerate(history, 2):
                        data = [h.get("일시", ""), h.get("구분", ""), h.get("유형", ""),
                                h.get("품번/제품코드", ""), h.get("품명", ""),
                                h.get("수량", ""), h.get("잔여재고", ""),
                                h.get("관련제품", ""), h.get("비고", "")]
                        for col, val in enumerate(data, 1):
                            cell = ws.cell(row=row, column=col, value=val)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center")

                    for col in range(1, len(headers) + 1):
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                    ws.column_dimensions["A"].width = 20

                if report_type == "all" and "Sheet" in wb.sheetnames:
                    del wb["Sheet"]

                wb.save(filepath)
                self.app.root.after(0, lambda: self.report_status.configure(
                    text=f"✅ 리포트가 저장되었습니다: {filepath}", fg=COLORS["success"]))
                self.app.root.after(0, lambda: messagebox.showinfo("완료", f"리포트가 저장되었습니다.\n{filepath}"))

            except Exception as e:
                err_msg = str(e)
                self.app.root.after(0, lambda: self.report_status.configure(
                    text=f"❌ 리포트 생성 실패: {err_msg}", fg=COLORS["danger"]))
                self.app.root.after(0, lambda: messagebox.showerror("오류", err_msg))

        threading.Thread(target=generate, daemon=True).start()
