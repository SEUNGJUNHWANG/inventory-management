"""
재고관리 시스템 - 입출고 이력 페이지
- 개별 이력 취소 (기존)
- 생산 일괄 취소: '생산입고' 이력 우클릭 시 제품 재고 + 소요 부품 재고를 한 번에 원복
"""

import tkinter as tk
from tkinter import ttk, messagebox
import threading
from datetime import datetime, timedelta
from core.constants import COLORS, FONT_FAMILY, FONT_SIZES, HISTORY_COLUMNS


def _bind_tree_scroll(tree):
    """Treeview에 마우스 휠 스크롤 바인딩 (hover 기반)"""
    def _on_mw(e):
        try:
            tree.yview_scroll(int(-1 * (e.delta / 120)), "units")
        except Exception:
            pass
    tree.bind("<Enter>", lambda e: tree.bind_all("<MouseWheel>", _on_mw))
    tree.bind("<Leave>", lambda e: tree.unbind_all("<MouseWheel>"))

class HistoryPage:
    def __init__(self, app):
        self.app = app
        self.history_tree = None
        self.history_menu = None
        self.hist_start = None
        self.hist_end = None
        self._loaded_rows = []   # 현재 로드된 전체 행 캐시 (검색 필터링용)
        self.search_var = None   # 검색어 StringVar

        # 탭 상태
        self._tab_loaded = {0: False, 1: False}
        self._tab_frames = []

        # 거래처별 집계 탭 상태
        self._sup_rows    = []   # get_supplier_io_summary() 원본 결과
        self._sup_months  = []   # 집계에 등장한 월 목록 (정렬)
        self._sup_tree    = None

    def render(self):
        # 페이지 재진입 시 lazy-load 상태 초기화 (이전 render의 잔류 상태 제거)
        self._tab_loaded = {0: False, 1: False}

        scroll_frame = self.app._create_scrollable_frame()

        header = tk.Frame(scroll_frame, bg=COLORS["bg"])
        header.pack(fill=tk.X, padx=5, pady=(0, 10))
        tk.Label(header, text="📜 입출고 이력", bg=COLORS["bg"],
                 fg=COLORS["text"], font=(FONT_FAMILY, FONT_SIZES["title"], "bold")).pack(side=tk.LEFT)

        style = ttk.Style()
        style.configure("History.TNotebook.Tab",
                        font=(FONT_FAMILY, FONT_SIZES["small"], "bold"),
                        padding=[14, 6])
        nb = ttk.Notebook(scroll_frame, style="History.TNotebook")
        nb.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 10))

        self._tab_frames = []
        for name in ("📜 상세 이력", "🏢 거래처별 집계"):
            f = tk.Frame(nb, bg=COLORS["bg"])
            nb.add(f, text=name)
            self._tab_frames.append(f)

        nb.bind("<<NotebookTabChanged>>",
                lambda e: self._on_tab_change(nb.index(nb.select())))
        self._load_tab(0)

    def _on_tab_change(self, idx):
        if not self._tab_loaded.get(idx):
            self._load_tab(idx)

    def _load_tab(self, idx):
        self._tab_loaded[idx] = True
        (self._build_detail_tab, self._build_supplier_tab)[idx](self._tab_frames[idx])

    # ═══════════════════════════════════════════════════════════════════
    # 탭 1 — 상세 이력
    # ═══════════════════════════════════════════════════════════════════
    def _build_detail_tab(self, scroll_frame):
        # 필터
        filter_frame = tk.Frame(scroll_frame, bg=COLORS["bg"])
        filter_frame.pack(fill=tk.X, padx=5, pady=(0, 10))

        tk.Label(filter_frame, text="기간:", bg=COLORS["bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self.hist_start = tk.Entry(filter_frame, font=(FONT_FAMILY, FONT_SIZES["small"]), width=12)
        self.hist_start.pack(side=tk.LEFT, padx=3)
        self.hist_start.insert(0, (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
        tk.Label(filter_frame, text="~", bg=COLORS["bg"]).pack(side=tk.LEFT)
        self.hist_end = tk.Entry(filter_frame, font=(FONT_FAMILY, FONT_SIZES["small"]), width=12)
        self.hist_end.pack(side=tk.LEFT, padx=3)
        self.hist_end.insert(0, datetime.now().strftime("%Y-%m-%d"))

        tk.Button(filter_frame, text="조회", font=(FONT_FAMILY, 9),
                  command=self._load_data).pack(side=tk.LEFT, padx=5)
        tk.Button(filter_frame, text="전체 조회", font=(FONT_FAMILY, 9),
                  command=lambda: self._load_data(all_data=True)).pack(side=tk.LEFT)

        # ── 검색 입력창 ──
        tk.Frame(filter_frame, bg=COLORS["bg"], width=20).pack(side=tk.LEFT)  # 구분 여백
        tk.Label(filter_frame, text="🔍 검색:", bg=COLORS["bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(filter_frame, textvariable=self.search_var,
                                font=(FONT_FAMILY, FONT_SIZES["small"]), width=20)
        search_entry.pack(side=tk.LEFT, padx=(3, 2))
        search_entry.bind("<KeyRelease>", lambda e: self._apply_search())
        search_entry.bind("<Control-a>", lambda e: (search_entry.select_range(0, tk.END), "break")[1])

        tk.Button(filter_frame, text="✕", font=(FONT_FAMILY, 8),
                  padx=4, pady=1, relief="flat",
                  command=lambda: (self.search_var.set(""), self._apply_search())).pack(side=tk.LEFT)

        self._count_label = tk.Label(filter_frame, text="", bg=COLORS["bg"],
                                     fg=COLORS["text_secondary"],
                                     font=(FONT_FAMILY, FONT_SIZES["small"]))
        self._count_label.pack(side=tk.LEFT, padx=(8, 0))

        card = tk.Frame(scroll_frame, bg=COLORS["card_bg"],
                        highlightbackground=COLORS["border"], highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=5)

        self.history_tree = ttk.Treeview(card, columns=HISTORY_COLUMNS, show="headings", height=20)
        for col in HISTORY_COLUMNS:
            self.history_tree.heading(col, text=col)
            self.history_tree.column(col, width=100, anchor="center")
        self.history_tree.column("No", width=50)
        self.history_tree.column("일시", width=155)
        self.history_tree.column("품명", width=200)
        self.history_tree.column("비고", width=150)

        # 반응형: 품명/비고 컬럼 너비 자동 조정
        def _on_hist_resize(event):
            total = self.history_tree.winfo_width()
            # No(50) + 일시(155) + 유형(100) + 수량(100) + 단위(80) + 스크롤바(20)
            fixed = 50 + 155 + 100 + 100 + 80 + 20
            remaining = max(200, total - fixed)
            part_w = int(remaining * 0.55)
            note_w = remaining - part_w
            self.history_tree.column("품명", width=part_w)
            self.history_tree.column("비고", width=note_w)
        self.history_tree.bind("<Configure>", _on_hist_resize)

        # 행 색상 태그 설정
        self.history_tree.tag_configure("생산입고", background="#dbeafe", foreground="#1e40af")
        self.history_tree.tag_configure("생산출고", background="#fef3c7", foreground="#92400e")
        self.history_tree.tag_configure("취소", background="#fecaca", foreground="#991b1b")

        hist_scroll = ttk.Scrollbar(card, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=hist_scroll.set)
        self.history_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        hist_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 우클릭 메뉴 (동적으로 구성)
        self.history_menu = tk.Menu(self.app.root, tearoff=0)
        self.history_tree.bind("<Button-3>", self._right_click)

        self._load_data()

    def _load_data(self, all_data=False):
        def load():
            try:
                if all_data:
                    history = self.app.db.get_all_history()
                else:
                    start = self.hist_start.get().strip()
                    end = self.hist_end.get().strip()
                    history = self.app.db.get_history_by_date_range(start, end)
                history.reverse()
                self.app.root.after(0, lambda: render(history, all_data))
            except Exception as e:
                err_msg = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err_msg))

        def render(history, is_all):
            all_hist = self.app.db.get_all_history() if not is_all else history[::-1]
            self._loaded_rows = []
            for idx, h in enumerate(history):
                row_no = len(all_hist) - idx + 1 if not is_all else len(history) - idx + 1
                h_type = h.get("유형", "")
                h_direction = h.get("구분", "")

                tag = "normal"
                if h_type == "생산입고":
                    tag = "생산입고"
                elif h_type == "생산출고":
                    tag = "생산출고"
                elif h_direction == "취소":
                    tag = "취소"

                self._loaded_rows.append({
                    "values": (
                        row_no,
                        h.get("일시", ""), h.get("구분", ""), h_type,
                        h.get("품번/제품코드", ""), h.get("품명", ""),
                        h.get("수량", ""), h.get("잔여재고", ""),
                        h.get("관련제품", ""), h.get("비고", ""),
                    ),
                    "tag": tag,
                })
            self._apply_search()

        threading.Thread(target=load, daemon=True).start()

    def _apply_search(self, *_):
        """현재 로드된 행을 검색어로 클라이언트 필터링 후 트리 재렌더링."""
        keyword = (self.search_var.get().strip().lower()
                   if self.search_var else "")
        self.history_tree.delete(*self.history_tree.get_children())

        matched = 0
        for row in self._loaded_rows:
            if not keyword or any(keyword in str(v).lower() for v in row["values"]):
                self.history_tree.insert("", "end",
                                         values=row["values"],
                                         tags=(row["tag"],))
                matched += 1

        total = len(self._loaded_rows)
        if hasattr(self, "_count_label") and self._count_label:
            if keyword:
                self._count_label.config(
                    text=f"{matched:,} / {total:,}건",
                    fg=COLORS["primary"] if matched else COLORS["danger"],
                )
            else:
                self._count_label.config(
                    text=f"총 {total:,}건",
                    fg=COLORS["text_secondary"],
                )

    def _right_click(self, event):
        item = self.history_tree.identify_row(event.y)
        if not item:
            return
        self.history_tree.selection_set(item)

        values = self.history_tree.item(item)["values"]
        h_type = str(values[3])  # 유형 컬럼

        # 메뉴 초기화 후 동적 구성
        self.history_menu.delete(0, tk.END)

        if h_type == "생산입고":
            # 생산입고 이력: 일괄 취소 메뉴 표시
            self.history_menu.add_command(
                label="🏭 생산 전체 취소 (제품 + 부품 재고 일괄 원복)",
                command=self._cancel_production,
                font=(FONT_FAMILY, FONT_SIZES["small"], "bold"),
            )
            self.history_menu.add_separator()
            self.history_menu.add_command(
                label="이 이력만 취소 (제품 재고만 원복)",
                command=self._cancel_history,
            )
        else:
            # 기타 이력: 기존 개별 취소 메뉴
            self.history_menu.add_command(
                label="이 이력 취소 (원복)",
                command=self._cancel_history,
            )

        self.history_menu.post(event.x_root, event.y_root)

    def _cancel_production(self):
        """생산 일괄 취소 - 제품 재고 + 소요 부품 재고 한 번에 원복"""
        selected = self.history_tree.selection()
        if not selected:
            return
        values = self.history_tree.item(selected[0])["values"]
        row_no = int(values[0])
        h_type = str(values[3])

        if h_type != "생산입고":
            messagebox.showwarning("알림", "생산입고 이력만 일괄 취소할 수 있습니다.")
            return

        # 상세 확인 대화상자
        info = (
            f"일시: {values[1]}\n"
            f"제품: {values[5]} ({values[4]})\n"
            f"생산수량: {values[6]}개\n\n"
            f"이 생산 건을 전체 취소하시겠습니까?\n\n"
            f"다음 작업이 일괄 수행됩니다:\n"
            f"  1. 제품 재고 -{values[6]}개 원복\n"
            f"  2. 소요된 모든 부품 재고 원복\n"
            f"  3. 취소 이력 자동 기록"
        )

        if not messagebox.askyesno("생산 전체 취소", info, icon="warning"):
            return

        # 처리 중 표시
        self.history_tree.configure(cursor="watch")

        def process():
            try:
                success, msg, details = self.app.db.cancel_production(row_no)
                self.app.root.after(0, lambda: self._show_production_cancel_result(success, msg, details))
            except Exception as e:
                err_msg = str(e)
                self.app.root.after(0, lambda: self._show_production_cancel_result(False, f"오류 발생: {err_msg}", []))

        threading.Thread(target=process, daemon=True).start()

    def _show_production_cancel_result(self, success, msg, details):
        """생산 일괄 취소 결과 표시"""
        self.history_tree.configure(cursor="")

        if success:
            # 결과를 보여주는 상세 대화상자
            result_dialog = tk.Toplevel(self.app.root)
            result_dialog.title("생산 일괄 취소 완료")
            result_dialog.transient(self.app.root)
            result_dialog.grab_set()

            # 크기 및 위치 설정
            dw, dh = 600, 450
            self.app.center_dialog(result_dialog, dw, dh)
            result_dialog.resizable(True, True)

            # 성공 아이콘 및 제목
            header_frame = tk.Frame(result_dialog, bg="#f0fdf4", padx=15, pady=12)
            header_frame.pack(fill=tk.X)
            tk.Label(header_frame, text="✅ 생산 일괄 취소가 완료되었습니다.",
                     bg="#f0fdf4", fg="#166534",
                     font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")).pack(anchor="w")

            # 상세 결과 텍스트
            text_frame = tk.Frame(result_dialog)
            text_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)

            result_text = tk.Text(text_frame, font=(FONT_FAMILY, FONT_SIZES["small"]),
                                  wrap="word", state="normal")
            result_scroll = ttk.Scrollbar(text_frame, orient="vertical", command=result_text.yview)
            result_text.configure(yscrollcommand=result_scroll.set)
            result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            result_scroll.pack(side=tk.RIGHT, fill=tk.Y)

            result_text.insert("1.0", msg)
            result_text.configure(state="disabled")

            # 닫기 버튼
            btn_frame = tk.Frame(result_dialog, padx=15, pady=10)
            btn_frame.pack(fill=tk.X)
            tk.Button(btn_frame, text="확인", font=(FONT_FAMILY, FONT_SIZES["body"], "bold"),
                      bg=COLORS["primary"], fg="white", padx=25, pady=6,
                      cursor="hand2",
                      command=result_dialog.destroy).pack(side=tk.RIGHT)

            # 이력 목록 새로고침
            self._load_data()
        else:
            messagebox.showerror("생산 취소 실패", msg)

    def _cancel_history(self):
        """기존 개별 이력 취소"""
        selected = self.history_tree.selection()
        if not selected:
            return
        values = self.history_tree.item(selected[0])["values"]
        row_no = int(values[0])

        info = f"일시: {values[1]}\n유형: {values[3]}\n품명: {values[5]}\n수량: {values[6]}"
        if not messagebox.askyesno("이력 취소", f"다음 이력을 취소하시겠습니까?\n\n{info}\n\n재고가 원복됩니다."):
            return

        def process():
            success, msg = self.app.db.cancel_history(row_no)
            self.app.root.after(0, lambda: self._show_result(success, msg))

        threading.Thread(target=process, daemon=True).start()

    def _show_result(self, success, msg):
        if success:
            messagebox.showinfo("취소 완료", msg)
            self._load_data()
        else:
            messagebox.showerror("취소 실패", msg)

    # ═══════════════════════════════════════════════════════════════════
    # 탭 2 — 거래처(매입처)별 집계
    # ═══════════════════════════════════════════════════════════════════

    # 지표 라벨 → (집계 필드, 금액여부)
    _SUP_METRICS = {
        "입고수량": ("입고수량", False),
        "출고수량": ("출고수량", False),
        "입고금액": ("입고금액", True),
        "출고금액": ("출고금액", True),
        "순증감(입고-출고)": (None, False),   # 파생 지표
    }

    def _build_supplier_tab(self, parent):
        # ── 필터 바 ────────────────────────────────────────────────────
        fbar = tk.Frame(parent, bg=COLORS["bg"])
        fbar.pack(fill=tk.X, padx=5, pady=(0, 8))

        tk.Label(fbar, text="기간:", bg=COLORS["bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._sup_start = tk.Entry(fbar, font=(FONT_FAMILY, FONT_SIZES["small"]), width=12)
        self._sup_start.pack(side=tk.LEFT, padx=3)
        self._sup_start.insert(0, (datetime.now() - timedelta(days=180)).strftime("%Y-%m-%d"))
        tk.Label(fbar, text="~", bg=COLORS["bg"]).pack(side=tk.LEFT)
        self._sup_end = tk.Entry(fbar, font=(FONT_FAMILY, FONT_SIZES["small"]), width=12)
        self._sup_end.pack(side=tk.LEFT, padx=3)
        self._sup_end.insert(0, datetime.now().strftime("%Y-%m-%d"))

        tk.Label(fbar, text="  업체:", bg=COLORS["bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._sup_name_var = tk.StringVar(value="전체")
        self._sup_name_cb = ttk.Combobox(fbar, textvariable=self._sup_name_var,
                                         values=["전체"], width=16, state="readonly")
        self._sup_name_cb.pack(side=tk.LEFT, padx=3)
        self._sup_name_cb.bind("<<ComboboxSelected>>", lambda e: self._sup_render())

        tk.Label(fbar, text="  지표:", bg=COLORS["bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._sup_metric_var = tk.StringVar(value="입고수량")
        metric_cb = ttk.Combobox(fbar, textvariable=self._sup_metric_var,
                                 values=list(self._SUP_METRICS.keys()),
                                 width=16, state="readonly")
        metric_cb.pack(side=tk.LEFT, padx=3)
        metric_cb.bind("<<ComboboxSelected>>", lambda e: self._sup_render())

        tk.Button(fbar, text="🔍 조회", font=(FONT_FAMILY, FONT_SIZES["small"], "bold"),
                  bg=COLORS["primary"], fg="white", padx=12, pady=3,
                  cursor="hand2", relief="flat",
                  command=self._sup_load).pack(side=tk.LEFT, padx=(8, 4))
        tk.Button(fbar, text="📥 엑셀", font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg=COLORS["success"], fg="white", padx=10, pady=3,
                  cursor="hand2", relief="flat",
                  command=self._sup_export_excel).pack(side=tk.LEFT)

        tk.Label(parent, text="※ 거래처는 부품마스터의 업체명 기준이며, 금액은 현재 단가 × 수량 추산값입니다. "
                              "취소 이력은 자동 상계됩니다.",
                 bg=COLORS["bg"], fg=COLORS["text_secondary"],
                 font=(FONT_FAMILY, FONT_SIZES["tiny"])).pack(anchor="w", padx=8, pady=(0, 6))

        # ── 요약 카드 ──────────────────────────────────────────────────
        self._sup_card_frame = tk.Frame(parent, bg=COLORS["bg"])
        self._sup_card_frame.pack(fill=tk.X, padx=5, pady=(0, 8))

        # ── 차트 ───────────────────────────────────────────────────────
        chart_outer = tk.Frame(parent, bg=COLORS["card_bg"],
                               highlightbackground=COLORS["border"], highlightthickness=1)
        chart_outer.pack(fill=tk.X, padx=5, pady=(0, 8))
        self._sup_chart_title = tk.Label(chart_outer, text="업체별 상위 10",
                                         bg=COLORS["card_bg"], fg=COLORS["text"],
                                         font=(FONT_FAMILY, FONT_SIZES["heading"], "bold"))
        self._sup_chart_title.pack(anchor="w", padx=15, pady=(10, 4))
        self._sup_chart_frame = tk.Frame(chart_outer, bg=COLORS["card_bg"])
        self._sup_chart_frame.pack(fill=tk.X, padx=15, pady=(0, 12))

        # ── 크로스탭 ───────────────────────────────────────────────────
        tbl_card = tk.Frame(parent, bg=COLORS["card_bg"],
                            highlightbackground=COLORS["border"], highlightthickness=1)
        tbl_card.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 10))

        hdr = tk.Frame(tbl_card, bg=COLORS["card_bg"])
        hdr.pack(fill=tk.X, padx=15, pady=(10, 4))
        tk.Label(hdr, text="거래처 × 월 집계", bg=COLORS["card_bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")).pack(side=tk.LEFT)
        tk.Label(hdr, text="(행을 더블클릭하면 해당 업체의 상세 이력을 볼 수 있습니다)",
                 bg=COLORS["card_bg"], fg=COLORS["text_secondary"],
                 font=(FONT_FAMILY, FONT_SIZES["tiny"])).pack(side=tk.LEFT, padx=8)

        tf = tk.Frame(tbl_card, bg=COLORS["card_bg"])
        tf.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 12))

        self._sup_tree = ttk.Treeview(tf, show="headings", height=14)
        vsb = ttk.Scrollbar(tf, orient="vertical", command=self._sup_tree.yview)
        hsb = ttk.Scrollbar(tf, orient="horizontal", command=self._sup_tree.xview)
        self._sup_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._sup_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)

        self._sup_tree.tag_configure("total", background="#f1f5f9",
                                     font=(FONT_FAMILY, FONT_SIZES["small"], "bold"))
        self._sup_tree.tag_configure("neg", foreground="#dc2626")
        self._sup_tree.bind("<Double-1>", self._sup_drilldown)

        self._sup_load()

    def _sup_load(self):
        start = self._sup_start.get().strip() or None
        end   = self._sup_end.get().strip() or None

        def load():
            try:
                rows      = self.app.db.get_supplier_io_summary(start, end)
                suppliers = sorted({r["업체명"] for r in rows})
                self.app.root.after(0, lambda: self._sup_on_loaded(rows, suppliers))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err))

        threading.Thread(target=load, daemon=True).start()

    def _sup_on_loaded(self, rows, suppliers):
        self._sup_rows = rows
        cur = self._sup_name_var.get()
        self._sup_name_cb.configure(values=["전체"] + suppliers)
        if cur != "전체" and cur not in suppliers:
            self._sup_name_var.set("전체")
        self._sup_render()

    def _sup_metric_value(self, row):
        """현재 선택된 지표에 해당하는 값을 행에서 추출."""
        field, _ = self._SUP_METRICS[self._sup_metric_var.get()]
        if field is None:                      # 순증감
            return row["입고수량"] - row["출고수량"]
        return row[field]

    def _sup_fmt(self, v):
        _, is_money = self._SUP_METRICS[self._sup_metric_var.get()]
        if v == 0:
            return "-"
        return f"{v:,}원" if is_money else f"{v:,}"

    def _sup_render(self):
        rows   = self._sup_rows
        picked = self._sup_name_var.get()
        if picked != "전체":
            rows = [r for r in rows if r["업체명"] == picked]

        months = sorted({r["월"] for r in rows})
        self._sup_months = months

        # 업체 → {월: 지표값}
        pivot: dict = {}
        for r in rows:
            pivot.setdefault(r["업체명"], {})[r["월"]] = self._sup_metric_value(r)

        # ── 컬럼 재구성 ────────────────────────────────────────────────
        cols = ["업체명"] + months + ["합계"]
        self._sup_tree.configure(columns=cols)
        for c in cols:
            self._sup_tree.heading(c, text=c)
            if c == "업체명":
                self._sup_tree.column(c, width=150, anchor="w", stretch=False)
            elif c == "합계":
                self._sup_tree.column(c, width=110, anchor="e", stretch=False)
            else:
                self._sup_tree.column(c, width=95, anchor="e", stretch=False)

        self._sup_tree.delete(*self._sup_tree.get_children())

        # 합계 기준 내림차순 정렬
        totals = {s: sum(m.values()) for s, m in pivot.items()}
        ordered = sorted(pivot.keys(), key=lambda s: totals[s], reverse=True)

        for s in ordered:
            vals = [s] + [self._sup_fmt(pivot[s].get(m, 0)) for m in months] \
                       + [self._sup_fmt(totals[s])]
            self._sup_tree.insert("", "end", values=vals,
                                  tags=("neg",) if totals[s] < 0 else ())

        # 합계 행
        if ordered:
            col_totals = [sum(pivot[s].get(m, 0) for s in ordered) for m in months]
            grand      = sum(col_totals)
            self._sup_tree.insert("", "end", tags=("total",), values=(
                ["합계"] + [self._sup_fmt(v) for v in col_totals] + [self._sup_fmt(grand)]
            ))

        self._sup_render_cards(rows, len(pivot))
        self._sup_render_chart(totals, ordered)

    def _sup_render_cards(self, rows, supplier_count):
        for w in self._sup_card_frame.winfo_children():
            w.destroy()

        in_qty  = sum(r["입고수량"] for r in rows)
        out_qty = sum(r["출고수량"] for r in rows)
        in_amt  = sum(r["입고금액"] for r in rows)
        net     = in_qty - out_qty

        cards = [
            ("총 입고수량", f"{in_qty:,}", COLORS["primary"]),
            ("총 출고수량", f"{out_qty:,}", COLORS["warning"]),
            ("입고금액(추산)", f"{in_amt:,}원", COLORS["info"]),
            ("순증감", f"{net:+,}", COLORS["success"] if net >= 0 else COLORS["danger"]),
            ("거래 업체 수", f"{supplier_count}곳", COLORS["text_secondary"]),
        ]
        for label, val, color in cards:
            c = tk.Frame(self._sup_card_frame, bg="white",
                         highlightbackground=color, highlightthickness=2)
            c.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=3)
            tk.Label(c, text=label, bg="white", fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["small"])).pack(pady=(8, 2))
            tk.Label(c, text=val, bg="white", fg=color,
                     font=(FONT_FAMILY, FONT_SIZES["stat"], "bold")).pack(pady=(0, 8))

    def _sup_render_chart(self, totals, ordered):
        for w in self._sup_chart_frame.winfo_children():
            w.destroy()
        metric = self._sup_metric_var.get()
        self._sup_chart_title.configure(text=f"업체별 {metric} 상위 10")

        top = ordered[:10]
        if not top:
            tk.Label(self._sup_chart_frame, text="데이터 없음",
                     bg=COLORS["card_bg"], fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["small"])).pack(pady=25)
            return

        max_abs = max((abs(totals[s]) for s in top), default=1) or 1
        BAR_MAX = 340

        for s in top:
            v = totals[s]
            row = tk.Frame(self._sup_chart_frame, bg=COLORS["card_bg"])
            row.pack(fill=tk.X, pady=1)

            tk.Label(row, text=s[:14], bg=COLORS["card_bg"], fg=COLORS["text"],
                     font=(FONT_FAMILY, FONT_SIZES["small"]),
                     width=14, anchor="w").pack(side=tk.LEFT)

            canvas = tk.Canvas(row, width=BAR_MAX + 6, height=20,
                               bg=COLORS["card_bg"], highlightthickness=0)
            canvas.pack(side=tk.LEFT)
            w = max(2, int(BAR_MAX * abs(v) / max_abs))
            canvas.create_rectangle(0, 4, w, 17,
                                    fill="#3b82f6" if v >= 0 else "#ef4444",
                                    outline="")

            tk.Label(row, text=self._sup_fmt(v), bg=COLORS["card_bg"],
                     fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["small"]),
                     anchor="w").pack(side=tk.LEFT, padx=(6, 0))

    def _sup_drilldown(self, event):
        """업체 행 더블클릭 → 해당 업체의 기간 내 상세 이력 팝업."""
        sel = self._sup_tree.selection()
        if not sel:
            return
        supplier = str(self._sup_tree.item(sel[0])["values"][0])
        if supplier == "합계":
            return

        start = self._sup_start.get().strip() or None
        end   = self._sup_end.get().strip() or None

        dlg = tk.Toplevel(self.app.root)
        dlg.title(f"상세 이력 - {supplier}")
        dlg.transient(self.app.root)
        self.app.center_dialog(dlg, 980, 560)

        tk.Label(dlg, text=f"🏢 {supplier}", fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")).pack(anchor="w", padx=15, pady=(12, 2))
        info_lbl = tk.Label(dlg, text="불러오는 중...", fg=COLORS["text_secondary"],
                            font=(FONT_FAMILY, FONT_SIZES["small"]))
        info_lbl.pack(anchor="w", padx=15, pady=(0, 8))

        cols = ("일시", "구분", "유형", "품번", "품명", "수량", "잔여재고", "비고")
        tree = ttk.Treeview(dlg, columns=cols, show="headings")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=110, anchor="center")
        tree.column("일시", width=150)
        tree.column("품명", width=180, anchor="w")
        tree.column("비고", width=150, anchor="w")
        vsb = ttk.Scrollbar(dlg, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(15, 0), pady=(0, 15))
        vsb.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 15), pady=(0, 15))
        tree.tag_configure("취소", background="#fecaca", foreground="#991b1b")

        def load():
            try:
                history   = self.app.db.get_all_history()
                parts_map = self.app.db._get_parts_map()
                types     = self.app.db._SUPPLIER_IO_TYPES
                nolabel   = self.app.db.NO_SUPPLIER_LABEL
                matched   = []
                for h in history:
                    if str(h.get("유형", "")) not in types:
                        continue
                    d = str(h.get("일시", ""))[:10]
                    if (start and d < start) or (end and d > end):
                        continue
                    part = parts_map.get(str(h.get("품번/제품코드", "")), {})
                    name = str(part.get("업체명", "")).strip() or nolabel
                    if name == supplier:
                        matched.append(h)
                matched.reverse()
                self.app.root.after(0, lambda: fill(matched))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err, parent=dlg))

        def fill(rows):
            for h in rows:
                tag = "취소" if str(h.get("구분", "")) == "취소" else ""
                tree.insert("", "end", values=(
                    h.get("일시", ""), h.get("구분", ""), h.get("유형", ""),
                    h.get("품번/제품코드", ""), h.get("품명", ""),
                    h.get("수량", ""), h.get("잔여재고", ""), h.get("비고", ""),
                ), tags=(tag,) if tag else ())
            period = f"{start or '전체'} ~ {end or '전체'}"
            info_lbl.configure(text=f"기간: {period}   |   총 {len(rows):,}건")

        threading.Thread(target=load, daemon=True).start()

    def _sup_export_excel(self):
        if not self._sup_tree or not self._sup_tree.get_children():
            messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
            return

        from tkinter import filedialog
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            title="거래처별 입출고 집계 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=f"거래처별_입출고집계_{now_str}.xlsx",
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            hdr_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
            nrm_font = Font(name="맑은 고딕", size=10)
            tot_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"), bottom=Side(style="thin"))

            wb = openpyxl.Workbook()

            # 시트 1 — 크로스탭 (화면 그대로)
            ws = wb.active
            ws.title = "거래처별 집계"
            metric = self._sup_metric_var.get()
            ws["A1"] = (f"거래처별 입출고 집계 — {metric}  "
                        f"({self._sup_start.get().strip()} ~ {self._sup_end.get().strip()})")
            ws["A1"].font = Font(name="맑은 고딕", bold=True, size=13)

            cols = list(self._sup_tree["columns"])
            for c, h in enumerate(cols, 1):
                cell = ws.cell(row=3, column=c, value=h)
                cell.font, cell.fill = hdr_font, hdr_fill
                cell.alignment, cell.border = Alignment(horizontal="center"), thin

            for i, iid in enumerate(self._sup_tree.get_children(), 4):
                vals    = self._sup_tree.item(iid)["values"]
                is_tot  = "total" in self._sup_tree.item(iid)["tags"]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=i, column=c, value=v)
                    cell.font = Font(name="맑은 고딕", bold=True, size=10) if is_tot else nrm_font
                    cell.border = thin
                    cell.alignment = Alignment(horizontal="left" if c == 1 else "right")
                    if is_tot:
                        cell.fill = tot_fill
            ws.column_dimensions["A"].width = 20
            for c in range(2, len(cols) + 1):
                ws.column_dimensions[get_column_letter(c)].width = 14

            # 시트 2 — 원본 집계 (업체 × 월 전 지표)
            ws2 = wb.create_sheet("원본 집계")
            headers2 = ["업체명", "월", "입고건수", "입고수량", "입고금액",
                        "출고건수", "출고수량", "출고금액", "순증감"]
            for c, h in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=c, value=h)
                cell.font, cell.fill = hdr_font, hdr_fill
                cell.alignment, cell.border = Alignment(horizontal="center"), thin

            picked = self._sup_name_var.get()
            src = self._sup_rows if picked == "전체" else \
                  [r for r in self._sup_rows if r["업체명"] == picked]
            for i, r in enumerate(src, 2):
                vals = [r["업체명"], r["월"], r["입고건수"], r["입고수량"], r["입고금액"],
                        r["출고건수"], r["출고수량"], r["출고금액"],
                        r["입고수량"] - r["출고수량"]]
                for c, v in enumerate(vals, 1):
                    cell = ws2.cell(row=i, column=c, value=v)
                    cell.font, cell.border = nrm_font, thin
                    cell.alignment = Alignment(horizontal="left" if c == 1 else "center")
            for c, w in enumerate([20, 12, 12, 12, 14, 12, 12, 14, 12], 1):
                ws2.column_dimensions[get_column_letter(c)].width = w

            wb.save(path)
            messagebox.showinfo("저장 완료", f"저장되었습니다.\n{path}")
            import os
            os.startfile(path)
        except Exception as e:
            messagebox.showerror("오류", str(e))
