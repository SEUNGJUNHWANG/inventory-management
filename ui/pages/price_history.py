# -*- coding: utf-8 -*-
"""
재고관리 시스템 - 단가변경이력 페이지
- 부품 탭: 부품 구매단가 변경이력 (기존)
- 제품 탭: 제품 판매단가 변경이력 (신규)
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
from datetime import datetime
from collections import defaultdict
from core.constants import COLORS, FONT_FAMILY, FONT_SIZES


class PriceHistoryPage:
    """단가변경이력 페이지 (부품/제품 탭)"""

    def __init__(self, app):
        self.app = app
        # 부품 탭 데이터
        self._part_records  = []
        self._part_suppliers = []
        # 제품 탭 데이터
        self._prod_records  = []

    # ═══════════════════════════════════════════
    # 렌더링
    # ═══════════════════════════════════════════
    def render(self):
        outer = self.app._create_scrollable_frame()

        # 타이틀 행
        title_f = tk.Frame(outer, bg=COLORS["bg"])
        title_f.pack(fill=tk.X, padx=5, pady=(5, 8))
        tk.Label(title_f, text="💰 단가변경이력",
                 bg=COLORS["bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["title"], "bold")).pack(side=tk.LEFT)
        tk.Button(title_f, text="🔄 새로고침",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg="#e2e8f0", fg=COLORS["text"], padx=10, pady=3,
                  cursor="hand2", command=self._load_all).pack(side=tk.RIGHT)

        # ── Notebook 탭 ──
        style = ttk.Style()
        style.configure("PH.TNotebook",        background=COLORS["bg"], borderwidth=0)
        style.configure("PH.TNotebook.Tab",    font=(FONT_FAMILY, FONT_SIZES["small"], "bold"),
                        padding=[18, 6])
        style.map("PH.TNotebook.Tab",
                  background=[("selected", COLORS["primary"]), ("!selected", "#e2e8f0")],
                  foreground=[("selected", "white"),            ("!selected", COLORS["text"])])

        nb = ttk.Notebook(outer, style="PH.TNotebook")
        nb.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 10))

        # 부품 탭
        part_frame = tk.Frame(nb, bg=COLORS["bg"])
        nb.add(part_frame, text="  🔩 부품 구매단가  ")
        self._build_part_tab(part_frame)

        # 제품 탭
        prod_frame = tk.Frame(nb, bg=COLORS["bg"])
        nb.add(prod_frame, text="  📦 제품 판매단가  ")
        self._build_prod_tab(prod_frame)

        # 탭 전환 시 해당 탭 데이터 로드
        nb.bind("<<NotebookTabChanged>>", self._on_tab_change)
        self._nb = nb

        # 초기 로드 (부품 탭)
        self._load_part_data()

    def _on_tab_change(self, event):
        idx = self._nb.index("current")
        if idx == 0 and not self._part_records:
            self._load_part_data()
        elif idx == 1 and not self._prod_records:
            self._load_prod_data()

    def _load_all(self):
        self._load_part_data()
        self._load_prod_data()

    # ═══════════════════════════════════════════
    # 공통 헬퍼
    # ═══════════════════════════════════════════
    def _make_dashboard(self, parent):
        """대시보드 카드 컨테이너 Frame 반환"""
        f = tk.Frame(parent, bg=COLORS["bg"])
        f.pack(fill=tk.X, padx=5, pady=(8, 6))
        return f

    def _make_chart_card(self, parent):
        f = tk.Frame(parent, bg=COLORS["card_bg"],
                     highlightbackground=COLORS["border"], highlightthickness=1)
        f.pack(fill=tk.X, padx=5, pady=(0, 6))
        return f

    def _render_dashboard(self, container, records):
        for w in container.winfo_children():
            w.destroy()
        now_ym     = datetime.now().strftime("%Y-%m")
        this_month = [r for r in records if r["변경일시"][:7] == now_ym]
        total_m    = len(this_month)
        up_m       = sum(1 for r in this_month if r["변경률"].startswith("+"))
        down_m     = sum(1 for r in this_month if r["변경률"].startswith("-"))
        rates = []
        for r in this_month:
            try:
                rates.append(float(r["변경률"].replace("%", "").replace("+", "")))
            except Exception:
                pass
        avg_rate = sum(rates) / len(rates) if rates else 0
        avg_str  = f"{avg_rate:+.1f}%" if rates else "-"

        stats = [
            ("이번달 변경",  f"{total_m}건",  COLORS["primary"]),
            ("평균 변동률",  avg_str,          COLORS["warning"] if avg_rate >= 0 else COLORS["success"]),
            ("인상 건수",    f"{up_m}건 🔴",   COLORS["danger"]),
            ("인하 건수",    f"{down_m}건 🟢", COLORS["success"]),
        ]
        for label, value, color in stats:
            c = tk.Frame(container, bg="white",
                         highlightbackground=color, highlightthickness=2)
            c.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=3)
            tk.Label(c, text=label, bg="white", fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["small"])).pack(pady=(10, 2))
            tk.Label(c, text=value, bg="white", fg=color,
                     font=(FONT_FAMILY, FONT_SIZES["stat"], "bold")).pack(pady=(0, 10))

    def _render_chart(self, chart_card, monthly, title="월별 단가 변경 현황 (최근 6개월)"):
        for w in chart_card.winfo_children():
            w.destroy()
        hdr = tk.Frame(chart_card, bg=COLORS["card_bg"])
        hdr.pack(fill=tk.X, padx=15, pady=(10, 4))
        tk.Label(hdr, text=title, bg=COLORS["card_bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")).pack(side=tk.LEFT)

        now = datetime.now()
        months_6 = []
        for i in range(5, -1, -1):
            m = now.month - i
            y = now.year
            while m <= 0:
                m += 12
                y -= 1
            months_6.append(f"{y}-{str(m).zfill(2)}")

        chart_frame = tk.Frame(chart_card, bg=COLORS["card_bg"])
        chart_frame.pack(fill=tk.X, padx=15, pady=(0, 14))

        max_val = max((monthly.get(ym, {}).get("total", 0) for ym in months_6), default=1)
        if max_val == 0:
            max_val = 1

        BAR_W, BAR_MAX_H, CANVAS_H = 60, 100, 150

        for ym in months_6:
            data   = monthly.get(ym, {"total": 0, "up": 0, "down": 0})
            total  = data["total"]
            up_cnt = data["up"]
            dn_cnt = data["down"]

            col_f = tk.Frame(chart_frame, bg=COLORS["card_bg"])
            col_f.pack(side=tk.LEFT, expand=True)

            canvas = tk.Canvas(col_f, width=BAR_W, height=CANVAS_H,
                                bg=COLORS["card_bg"], highlightthickness=0)
            canvas.pack()

            bar_h = int(BAR_MAX_H * total / max_val) if total > 0 else 2
            bot_y = CANVAS_H - 30

            if total > 0:
                up_h  = int(bar_h * up_cnt  / total)
                dn_h  = int(bar_h * dn_cnt  / total)
                etc_h = bar_h - up_h - dn_h
                x0, x1, y_cur = BAR_W // 2 - 14, BAR_W // 2 + 14, bot_y
                for h, color in [(dn_h, "#22c55e"), (etc_h, "#6366f1"), (up_h, "#ef4444")]:
                    if h > 0:
                        canvas.create_rectangle(x0, y_cur - h, x1, y_cur,
                                                fill=color, outline="")
                        y_cur -= h
                canvas.create_text(BAR_W // 2, bot_y - bar_h - 6,
                                   text=str(total), font=(FONT_FAMILY, 9, "bold"),
                                   fill=COLORS["text"])
            else:
                canvas.create_rectangle(BAR_W // 2 - 14, bot_y - 2, BAR_W // 2 + 14, bot_y,
                                        fill=COLORS["border"], outline="")

            canvas.create_text(BAR_W // 2, CANVAS_H - 14,
                                text=ym[5:] + "월", font=(FONT_FAMILY, 9),
                                fill=COLORS["text_secondary"])

        legend = tk.Frame(chart_card, bg=COLORS["card_bg"])
        legend.pack(pady=(0, 10))
        for color, text in [("#ef4444", "인상"), ("#22c55e", "인하"), ("#6366f1", "신규")]:
            tk.Frame(legend, bg=color, width=12, height=12).pack(side=tk.LEFT, padx=(8, 2))
            tk.Label(legend, text=text, bg=COLORS["card_bg"], fg=COLORS["text_secondary"],
                     font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT, padx=(0, 6))

    def _make_table(self, parent, col_defs):
        """(cid, heading, width, anchor) 목록으로 Treeview + Scrollbar 생성"""
        card = tk.Frame(parent, bg=COLORS["card_bg"],
                        highlightbackground=COLORS["border"], highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 10))

        hdr = tk.Frame(card, bg=COLORS["card_bg"])
        hdr.pack(fill=tk.X, padx=15, pady=(10, 4))
        tk.Label(hdr, text="상세 변경 이력", bg=COLORS["card_bg"], fg=COLORS["text"],
                 font=(FONT_FAMILY, FONT_SIZES["heading"], "bold")).pack(side=tk.LEFT)
        count_lbl = tk.Label(hdr, text="(0건)", bg=COLORS["card_bg"],
                              fg=COLORS["text_secondary"],
                              font=(FONT_FAMILY, FONT_SIZES["small"]))
        count_lbl.pack(side=tk.LEFT, padx=8)

        tf   = tk.Frame(card, bg=COLORS["card_bg"])
        tf.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 12))
        cols = [c[0] for c in col_defs]
        tree = ttk.Treeview(tf, columns=cols, show="headings", height=16)
        for cid, heading, width, anchor in col_defs:
            tree.heading(cid, text=heading)
            tree.column(cid, width=width, anchor=anchor, minwidth=40,
                        stretch=(cid not in ("no",)))
        vsb = ttk.Scrollbar(tf, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.tag_configure("up",   background="#fef2f2", foreground="#dc2626")
        tree.tag_configure("down", background="#f0fdf4", foreground="#16a34a")
        tree.tag_configure("new",  background="#eff6ff", foreground="#2563eb")
        return tree, count_lbl

    def _fill_table(self, tree, count_lbl, rows, col_keys):
        tree.delete(*tree.get_children())
        count_lbl.configure(text=f"({len(rows)}건)")
        for i, r in enumerate(rows, 1):
            rate = r.get("변경률", "")
            tag  = "up" if rate.startswith("+") else ("down" if rate.startswith("-") else "new")
            vals = [i] + [r.get(k, "") for k in col_keys]
            # 단가 포맷
            for j, k in enumerate(col_keys):
                if "단가" in k or "판매가" in k:
                    v = r.get(k, 0)
                    vals[j + 1] = f"{float(v):,.0f}" if v else "-"
            tree.insert("", tk.END, values=vals, tags=(tag,))

    # ═══════════════════════════════════════════
    # ── 부품 탭 ──────────────────────────────
    # ═══════════════════════════════════════════
    def _build_part_tab(self, parent):
        self._part_dashboard = self._make_dashboard(parent)
        self._part_chart     = self._make_chart_card(parent)
        self._build_part_filter(parent)
        self._part_tree, self._part_count_lbl = self._make_table(parent, [
            ("no",         "No",     40,  "center"),
            ("dt",         "변경일시", 140, "center"),
            ("part_id",    "품번",    100, "center"),
            ("part_name",  "부품명",  160, "w"),
            ("supplier",   "업체명",   90, "w"),
            ("old_price",  "이전단가",  90, "center"),
            ("new_price",  "변경단가",  90, "center"),
            ("rate",       "변경률",   70, "center"),
            ("changed_by", "변경자",   70, "center"),
            ("reason",     "변경사유", 130, "w"),
        ])

    def _build_part_filter(self, parent):
        card = tk.Frame(parent, bg=COLORS["card_bg"],
                        highlightbackground=COLORS["border"], highlightthickness=1)
        card.pack(fill=tk.X, padx=5, pady=(0, 4))
        row = tk.Frame(card, bg=COLORS["card_bg"])
        row.pack(fill=tk.X, padx=15, pady=10)

        now = datetime.now()
        tk.Label(row, text="연도:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._p_year = tk.StringVar(value=str(now.year))
        ttk.Combobox(row, textvariable=self._p_year, width=7, state="readonly",
                     values=[str(y) for y in range(now.year, now.year - 5, -1)]
                     ).pack(side=tk.LEFT, padx=(4, 12))

        tk.Label(row, text="월:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._p_month = tk.StringVar(value="전체")
        ttk.Combobox(row, textvariable=self._p_month, width=6, state="readonly",
                     values=["전체"] + [str(m) for m in range(1, 13)]
                     ).pack(side=tk.LEFT, padx=(4, 12))

        tk.Label(row, text="업체:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._p_supplier = tk.StringVar(value="전체")
        self._p_supplier_cb = ttk.Combobox(row, textvariable=self._p_supplier,
                                            width=14, state="readonly", values=["전체"])
        self._p_supplier_cb.pack(side=tk.LEFT, padx=(4, 12))

        tk.Label(row, text="구분:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._p_dir = tk.StringVar(value="전체")
        ttk.Combobox(row, textvariable=self._p_dir, width=9, state="readonly",
                     values=["전체", "인상(+)", "인하(-)", "신규"]
                     ).pack(side=tk.LEFT, padx=(4, 16))

        tk.Button(row, text="🔍 조회",
                  font=(FONT_FAMILY, FONT_SIZES["small"], "bold"),
                  bg=COLORS["primary"], fg="white", padx=14, pady=3,
                  cursor="hand2", command=self._apply_part_filter
                  ).pack(side=tk.LEFT, padx=(0, 6))
        tk.Button(row, text="📥 엑셀 내보내기",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg=COLORS["success"], fg="white", padx=12, pady=3,
                  cursor="hand2",
                  command=lambda: self._export_excel(self._part_tree, "부품단가변경이력")
                  ).pack(side=tk.LEFT)

    def _load_part_data(self):
        def load():
            try:
                records = self.app.db.get_price_history()
                monthly = self.app.db.get_price_history_monthly_summary()
                self.app.root.after(0, lambda: self._on_part_loaded(records, monthly))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err))
        threading.Thread(target=load, daemon=True).start()

    def _on_part_loaded(self, records, monthly):
        self._part_records = records
        suppliers = sorted({r["업체명"] for r in records if r["업체명"]})
        self._p_supplier_cb.configure(values=["전체"] + suppliers)
        self._render_dashboard(self._part_dashboard, records)
        self._render_chart(self._part_chart, monthly)
        self._apply_part_filter()

    def _apply_part_filter(self):
        year     = self._p_year.get()
        month    = self._p_month.get()
        supplier = self._p_supplier.get()
        dir_sel  = self._p_dir.get()
        filtered = []
        for r in self._part_records:
            dt = r["변경일시"]
            if not dt.startswith(year):
                continue
            if month != "전체" and not dt.startswith(f"{year}-{month.zfill(2)}"):
                continue
            if supplier != "전체" and r["업체명"] != supplier:
                continue
            rate = r["변경률"]
            if dir_sel == "인상(+)" and not rate.startswith("+"):
                continue
            if dir_sel == "인하(-)" and not rate.startswith("-"):
                continue
            if dir_sel == "신규" and rate != "신규":
                continue
            filtered.append(r)
        filtered.sort(key=lambda x: x["변경일시"], reverse=True)
        # 테이블 직접 렌더
        self._part_tree.delete(*self._part_tree.get_children())
        self._part_count_lbl.configure(text=f"({len(filtered)}건)")
        for i, r in enumerate(filtered, 1):
            rate = r["변경률"]
            tag  = "up" if rate.startswith("+") else ("down" if rate.startswith("-") else "new")
            old_p = f"{r['이전단가']:,.0f}" if r['이전단가'] else "-"
            new_p = f"{r['변경단가']:,.0f}" if r['변경단가'] else "-"
            self._part_tree.insert("", tk.END, tags=(tag,), values=(
                i, r["변경일시"], r["품번"], r["부품명"], r["업체명"],
                old_p, new_p, rate, r["변경자"], r["변경사유"],
            ))

    # ═══════════════════════════════════════════
    # ── 제품 탭 ──────────────────────────────
    # ═══════════════════════════════════════════
    def _build_prod_tab(self, parent):
        self._prod_dashboard = self._make_dashboard(parent)
        self._prod_chart     = self._make_chart_card(parent)
        self._build_prod_filter(parent)
        self._prod_tree, self._prod_count_lbl = self._make_table(parent, [
            ("no",         "No",      40,  "center"),
            ("dt",         "변경일시",  140, "center"),
            ("prod_id",    "제품코드",  110, "center"),
            ("prod_name",  "제품명",   170, "w"),
            ("old_price",  "이전판매가", 100, "center"),
            ("new_price",  "변경판매가", 100, "center"),
            ("rate",       "변경률",    70, "center"),
            ("changed_by", "변경자",    70, "center"),
            ("reason",     "변경사유",  140, "w"),
        ])

    def _build_prod_filter(self, parent):
        card = tk.Frame(parent, bg=COLORS["card_bg"],
                        highlightbackground=COLORS["border"], highlightthickness=1)
        card.pack(fill=tk.X, padx=5, pady=(0, 4))
        row = tk.Frame(card, bg=COLORS["card_bg"])
        row.pack(fill=tk.X, padx=15, pady=10)

        now = datetime.now()
        tk.Label(row, text="연도:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._q_year = tk.StringVar(value=str(now.year))
        ttk.Combobox(row, textvariable=self._q_year, width=7, state="readonly",
                     values=[str(y) for y in range(now.year, now.year - 5, -1)]
                     ).pack(side=tk.LEFT, padx=(4, 12))

        tk.Label(row, text="월:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._q_month = tk.StringVar(value="전체")
        ttk.Combobox(row, textvariable=self._q_month, width=6, state="readonly",
                     values=["전체"] + [str(m) for m in range(1, 13)]
                     ).pack(side=tk.LEFT, padx=(4, 12))

        tk.Label(row, text="구분:", bg=COLORS["card_bg"],
                 font=(FONT_FAMILY, FONT_SIZES["small"])).pack(side=tk.LEFT)
        self._q_dir = tk.StringVar(value="전체")
        ttk.Combobox(row, textvariable=self._q_dir, width=9, state="readonly",
                     values=["전체", "인상(+)", "인하(-)", "신규"]
                     ).pack(side=tk.LEFT, padx=(4, 16))

        tk.Button(row, text="🔍 조회",
                  font=(FONT_FAMILY, FONT_SIZES["small"], "bold"),
                  bg=COLORS["primary"], fg="white", padx=14, pady=3,
                  cursor="hand2", command=self._apply_prod_filter
                  ).pack(side=tk.LEFT, padx=(0, 6))
        tk.Button(row, text="📥 엑셀 내보내기",
                  font=(FONT_FAMILY, FONT_SIZES["small"]),
                  bg=COLORS["success"], fg="white", padx=12, pady=3,
                  cursor="hand2",
                  command=lambda: self._export_excel(self._prod_tree, "제품판매단가이력")
                  ).pack(side=tk.LEFT)

    def _load_prod_data(self):
        def load():
            try:
                records = self.app.db.get_product_price_history()
                monthly = self.app.db.get_product_price_history_monthly_summary()
                self.app.root.after(0, lambda: self._on_prod_loaded(records, monthly))
            except Exception as e:
                err = str(e)
                self.app.root.after(0, lambda: messagebox.showerror("오류", err))
        threading.Thread(target=load, daemon=True).start()

    def _on_prod_loaded(self, records, monthly):
        self._prod_records = records
        self._render_dashboard(self._prod_dashboard, records)
        self._render_chart(self._prod_chart, monthly, "월별 판매단가 변경 현황 (최근 6개월)")
        self._apply_prod_filter()

    def _apply_prod_filter(self):
        year    = self._q_year.get()
        month   = self._q_month.get()
        dir_sel = self._q_dir.get()
        filtered = []
        for r in self._prod_records:
            dt = r["변경일시"]
            if not dt.startswith(year):
                continue
            if month != "전체" and not dt.startswith(f"{year}-{month.zfill(2)}"):
                continue
            rate = r["변경률"]
            if dir_sel == "인상(+)" and not rate.startswith("+"):
                continue
            if dir_sel == "인하(-)" and not rate.startswith("-"):
                continue
            if dir_sel == "신규" and rate != "신규":
                continue
            filtered.append(r)
        filtered.sort(key=lambda x: x["변경일시"], reverse=True)
        self._prod_tree.delete(*self._prod_tree.get_children())
        self._prod_count_lbl.configure(text=f"({len(filtered)}건)")
        for i, r in enumerate(filtered, 1):
            rate = r["변경률"]
            tag  = "up" if rate.startswith("+") else ("down" if rate.startswith("-") else "new")
            old_p = f"{r['이전판매가']:,.0f}" if r['이전판매가'] else "-"
            new_p = f"{r['변경판매가']:,.0f}" if r['변경판매가'] else "-"
            self._prod_tree.insert("", tk.END, tags=(tag,), values=(
                i, r["변경일시"], r["제품코드"], r["제품명"],
                old_p, new_p, rate, r["변경자"], r["변경사유"],
            ))

    # ═══════════════════════════════════════════
    # 엑셀 내보내기 (공통)
    # ═══════════════════════════════════════════
    def _export_excel(self, tree, sheet_name):
        rows = [tree.item(iid)["values"] for iid in tree.get_children()]
        if not rows:
            messagebox.showinfo("알림", "내보낼 데이터가 없습니다.")
            return

        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            title="저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=f"{sheet_name}_{now_str}.xlsx",
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            hdr_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
            nrm_font = Font(name="맑은 고딕", size=10)
            up_fill  = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
            dn_fill  = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
            nw_fill  = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            thin     = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"),  bottom=Side(style="thin"))

            # 헤더는 트리뷰 컬럼 텍스트에서 추출
            headers = [tree.heading(col)["text"] for col in tree["columns"]]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin

            for r_idx, row in enumerate(rows, 2):
                rate_val = str(row[7] if len(row) > 7 else "")
                fill = up_fill if rate_val.startswith("+") else (
                       dn_fill if rate_val.startswith("-") else nw_fill)
                for c_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = nrm_font
                    cell.fill = fill
                    cell.border = thin
                    cell.alignment = Alignment(horizontal="center")

            for i in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16

            wb.save(path)
            messagebox.showinfo("저장 완료", f"저장되었습니다.\n{path}")
            import os
            os.startfile(path)
        except Exception as e:
            messagebox.showerror("오류", str(e))
