"""
부품 대량 등록용 엑셀 양식 템플릿 생성
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys


def create_parts_template(output_path):
    """부품 대량 등록/업데이트용 엑셀 양식 생성"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "부품등록"

    # 스타일 정의
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    example_fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
    required_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

    # 제목 행
    title_font = Font(name="맑은 고딕", size=14, bold=True, color="1E40AF")
    ws.merge_cells("A1:G1")
    ws["A1"] = "부품 대량 등록/업데이트 양식"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 안내 행
    guide_font = Font(name="맑은 고딕", size=9, color="6B7280")
    ws.merge_cells("A2:G2")
    ws["A2"] = "※ 품번(필수)이 이미 존재하면 업데이트, 없으면 신규 등록됩니다.  |  빨간 배경 = 필수 항목  |  파란 배경 = 예시 데이터 (삭제 후 사용)"
    ws["A2"].font = guide_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 25

    # 헤더 (3행)
    headers = [
        ("품번 (필수)", 18),
        ("부품명 (필수)", 30),
        ("규격", 35),
        ("단위", 10),
        ("현재재고", 12),
        ("안전재고", 12),
        ("비고", 25),
    ]

    for col_idx, (header_text, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 30

    # 예시 데이터 (4~6행)
    examples = [
        ["11072CA", "KB130몸통 반제품", "카트리지,락링,커버", "EA", 100, 20, "수입오동철"],
        ["21124NO-C", "노출배관(크롬)", "Ø15x300mm", "EA", 50, 10, ""],
        ["52080CO", "편심커플링", "Ø15 크롬", "EA", 200, 30, "국내생산"],
    ]

    data_font = Font(name="맑은 고딕", size=10)
    data_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    for row_idx, example in enumerate(examples, 4):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.fill = example_fill
            cell.border = thin_border
            if col_idx in (1, 4, 5, 6):
                cell.alignment = data_align
            else:
                cell.alignment = left_align

    # 빈 입력 행 (7~106행, 100행 준비)
    for row_idx in range(7, 107):
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            cell.border = thin_border
            cell.font = data_font
            if col_idx in (1, 2):
                cell.fill = required_fill  # 필수 항목 표시
            if col_idx in (1, 4, 5, 6):
                cell.alignment = data_align
            else:
                cell.alignment = left_align

    # 필터 설정
    ws.auto_filter.ref = f"A3:G{106}"

    # 시트 보호 안내 시트
    ws2 = wb.create_sheet("작성 안내")
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 60

    guide_data = [
        ("항목", "설명"),
        ("품번 (필수)", "부품의 고유 식별 코드입니다. 중복되면 기존 부품 정보가 업데이트됩니다."),
        ("부품명 (필수)", "부품의 이름입니다."),
        ("규격", "재질, 규격, 사이즈 등 부품의 상세 정보입니다."),
        ("단위", "수량 단위입니다. (예: EA, SET, BOX, KG 등) 비워두면 'EA'로 자동 설정됩니다."),
        ("현재재고", "현재 보유 재고 수량입니다. 비워두면 0으로 설정됩니다."),
        ("안전재고", "안전재고 기준 수량입니다. 재고가 이 수량 이하로 떨어지면 경고가 표시됩니다."),
        ("비고", "업체명, 메모 등 자유롭게 입력할 수 있습니다."),
        ("", ""),
        ("주의사항", ""),
        ("1", "예시 데이터(파란 배경)는 삭제하거나 덮어쓴 후 사용하세요."),
        ("2", "품번과 부품명은 반드시 입력해야 합니다."),
        ("3", "품번이 기존에 등록된 부품과 동일하면 해당 부품의 정보가 업데이트됩니다."),
        ("4", "100행 이상 등록이 필요하면 빈 행을 추가하여 사용하세요."),
        ("5", "엑셀 파일을 저장한 후 앱의 '부품 관리 > 엑셀 대량 등록' 버튼으로 업로드하세요."),
    ]

    for row_idx, (col_a, col_b) in enumerate(guide_data, 1):
        ws2.cell(row=row_idx, column=1, value=col_a).font = Font(name="맑은 고딕", size=10, bold=(row_idx <= 1 or col_a == "주의사항"))
        ws2.cell(row=row_idx, column=2, value=col_b).font = Font(name="맑은 고딕", size=10)

    wb.save(output_path)
    print(f"템플릿 생성 완료: {output_path}")
    return output_path


if __name__ == "__main__":
    output = sys.argv[1] if len(sys.argv) > 1 else "부품_대량등록_양식.xlsx"
    create_parts_template(output)
