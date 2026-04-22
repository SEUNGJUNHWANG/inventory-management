"""
BOM 엑셀 파일 임포트 도구
- '테노바 2026년 BOM.xlsx' 파일에서 부품리스트와 제품별 BOM 데이터를 읽어
  구글 시트에 자동으로 업로드합니다.
"""

import openpyxl
import os
import sys
import time


def parse_bom_excel(filepath):
    """
    BOM 엑셀 파일을 파싱하여 부품 목록과 제품별 BOM을 반환합니다.
    
    Returns:
        parts: [{"코드번호": ..., "업체명": ..., "부품명": ..., "규격": ..., "이전단가": ..., "현재단가": ...}, ...]
        products: [{"제품코드": ..., "제품명": ...}, ...]
        bom_items: [{"제품코드": ..., "부품품번": ..., "소요량": ...}, ...]
    """
    print(f"파일 로드 중: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # 1. 부품리스트 파싱
    parts = []
    ws_parts = wb["부품리스트"]
    print(f"부품리스트 파싱 중... (총 {ws_parts.max_row}행)")
    
    for row in range(2, ws_parts.max_row + 1):
        code = ws_parts.cell(row=row, column=1).value
        if not code or str(code).strip() == "":
            continue
        
        parts.append({
            "코드번호": str(code).strip(),
            "업체명": str(ws_parts.cell(row=row, column=2).value or "").strip(),
            "부품명": str(ws_parts.cell(row=row, column=3).value or "").strip(),
            "규격": str(ws_parts.cell(row=row, column=4).value or "").strip(),
            "이전단가": ws_parts.cell(row=row, column=5).value or 0,
            "현재단가": ws_parts.cell(row=row, column=6).value or 0,
        })
    
    print(f"  → 부품 {len(parts)}건 파싱 완료")
    
    # 2. 제품별 BOM 파싱
    products = []
    bom_items = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name == "부품리스트":
            continue
        
        ws = wb[sheet_name]
        
        # 2행 A열에 제품명이 있음
        product_name = ws.cell(row=2, column=1).value
        if not product_name:
            continue
        
        product_name = str(product_name).strip()
        product_code = sheet_name.strip()
        
        products.append({
            "제품코드": product_code,
            "제품명": product_name,
        })
        
        # 4행부터 BOM 데이터 (A열: 코드번호, G열: 소요량)
        bom_count = 0
        for row in range(4, ws.max_row + 1):
            part_code = ws.cell(row=row, column=1).value
            qty = ws.cell(row=row, column=7).value  # G열 = 소요량
            
            if not part_code or str(part_code).strip() == "":
                continue
            
            try:
                qty_val = float(qty) if qty else 0
            except (ValueError, TypeError):
                qty_val = 0
            
            if qty_val > 0:
                bom_items.append({
                    "제품코드": product_code,
                    "부품품번": str(part_code).strip(),
                    "소요량": qty_val,
                })
                bom_count += 1
        
        # print(f"  {product_code} ({product_name}): BOM {bom_count}건")
    
    print(f"  → 제품 {len(products)}건, BOM {len(bom_items)}건 파싱 완료")
    
    return parts, products, bom_items


def upload_to_google_sheets(db, parts, products, bom_items, progress_callback=None):
    """
    파싱된 데이터를 구글 시트에 업로드합니다.
    """
    import gspread
    
    total_steps = len(parts) + len(products) + len(bom_items)
    current_step = 0
    
    # 1. 부품 업로드
    print("\n부품 데이터 업로드 중...")
    ws_parts = db.spreadsheet.worksheet("부품마스터")
    
    # 기존 데이터 확인
    existing_parts = {str(p["품번"]): p for p in db.get_all_parts()}
    
    batch_data = []
    for p in parts:
        code = p["코드번호"]
        if code not in existing_parts:
            batch_data.append([
                code,
                p["부품명"],
                p["규격"],
                "EA",  # 기본 단위
                0,     # 현재재고 (초기값 0)
                0,     # 안전재고 (초기값 0)
                p["업체명"],  # 비고에 업체명
            ])
        current_step += 1
        if progress_callback and current_step % 100 == 0:
            progress_callback(current_step, total_steps, f"부품 처리 중... {current_step}/{len(parts)}")
    
    if batch_data:
        # 배치 업로드 (API 호출 최소화)
        print(f"  → 신규 부품 {len(batch_data)}건 업로드 중...")
        # gspread batch update
        start_row = len(existing_parts) + 2
        chunk_size = 500
        for i in range(0, len(batch_data), chunk_size):
            chunk = batch_data[i:i+chunk_size]
            cell_range = f"A{start_row + i}:G{start_row + i + len(chunk) - 1}"
            ws_parts.update(cell_range, chunk)
            print(f"    업로드: {i+1}~{min(i+chunk_size, len(batch_data))} / {len(batch_data)}")
            time.sleep(1)  # API 제한 방지
    
    # 2. 제품 업로드
    print("\n제품 데이터 업로드 중...")
    ws_products = db.spreadsheet.worksheet("제품마스터")
    existing_products = {str(p["제품코드"]): p for p in db.get_all_products()}
    
    batch_data = []
    for p in products:
        code = p["제품코드"]
        if code not in existing_products:
            batch_data.append([
                code,
                p["제품명"],
                "",  # 규격
                0,   # 현재재고
                "",  # 비고
            ])
        current_step += 1
    
    if batch_data:
        print(f"  → 신규 제품 {len(batch_data)}건 업로드 중...")
        start_row = len(existing_products) + 2
        chunk_size = 500
        for i in range(0, len(batch_data), chunk_size):
            chunk = batch_data[i:i+chunk_size]
            cell_range = f"A{start_row + i}:E{start_row + i + len(chunk) - 1}"
            ws_products.update(cell_range, chunk)
            print(f"    업로드: {i+1}~{min(i+chunk_size, len(batch_data))} / {len(batch_data)}")
            time.sleep(1)
    
    # 3. BOM 업로드
    print("\nBOM 데이터 업로드 중...")
    ws_bom = db.spreadsheet.worksheet("BOM")
    existing_bom = db.get_all_bom()
    existing_bom_set = set()
    for b in existing_bom:
        existing_bom_set.add((str(b["제품코드"]), str(b["부품품번"])))
    
    batch_data = []
    for b in bom_items:
        key = (b["제품코드"], b["부품품번"])
        if key not in existing_bom_set:
            batch_data.append([
                b["제품코드"],
                b["부품품번"],
                b["소요량"],
                "",  # 비고
            ])
        current_step += 1
    
    if batch_data:
        print(f"  → 신규 BOM {len(batch_data)}건 업로드 중...")
        start_row = len(existing_bom) + 2
        chunk_size = 500
        for i in range(0, len(batch_data), chunk_size):
            chunk = batch_data[i:i+chunk_size]
            cell_range = f"A{start_row + i}:D{start_row + i + len(chunk) - 1}"
            ws_bom.update(cell_range, chunk)
            print(f"    업로드: {i+1}~{min(i+chunk_size, len(batch_data))} / {len(batch_data)}")
            time.sleep(1)
    
    print("\n✅ 데이터 임포트 완료!")
    return len(batch_data)


if __name__ == "__main__":
    # 테스트용 독립 실행
    if len(sys.argv) < 2:
        print("사용법: python import_bom.py <BOM엑셀파일경로>")
        sys.exit(1)
    
    filepath = sys.argv[1]
    parts, products, bom_items = parse_bom_excel(filepath)
    print(f"\n파싱 결과:")
    print(f"  부품: {len(parts)}건")
    print(f"  제품: {len(products)}건")
    print(f"  BOM: {len(bom_items)}건")
